"""
Spreadsheet parser for AQA Combined Science Trilogy Scheme of Work.

Reads the .xlsx file, extracts lesson data from Year 10 and Year 11 sheets,
filters out assessment/review/DIRT rows, and produces structured lesson dicts
matching the booklet production schema.
"""

import re
from pathlib import Path
from openpyxl import load_workbook


# Hard non-booklet: these ALWAYS exclude regardless of spec content
HARD_NON_BOOKLET_PATTERNS = [
    r"\bassessment\b",
    r"\bdirt\b",
    r"\bmock\b",
    r"\brevision\b",
    r"\bexam\s+practice\b",
    r"\bwalking\s+talking\b",
    r"\bcontingency\b",
    r"\bexam\s+readiness\b",
]

# Soft non-booklet: only exclude if spec_content is missing or purely consolidation
SOFT_NON_BOOKLET_PATTERNS = [
    r"\breview\b",
    r"\bconsolidation\b",
    r"\brecap\b",
    r"\btopic\s+review\b",
    r"\bpractice\b",
]

HARD_NON_BOOKLET_RE = re.compile(
    "|".join(HARD_NON_BOOKLET_PATTERNS), re.IGNORECASE
)
SOFT_NON_BOOKLET_RE = re.compile(
    "|".join(SOFT_NON_BOOKLET_PATTERNS), re.IGNORECASE
)

# Spec content that indicates consolidation rather than new teaching
CONSOLIDATION_SPEC_RE = re.compile(
    r"^(consolidation|timed assessment|review|dedicated improvement|"
    r"full paper|teacher-led walkthrough|retrieval|rate calculations from graphs|"
    r"crude oil.*practice equations|all rps|review papers)",
    re.IGNORECASE,
)

# Column mapping (0-indexed) based on the header row
COL_MAP = {
    "week": 0,
    "lesson_num": 1,
    "subject": 2,
    "topic": 3,
    "title": 4,
    "spec_content": 5,
    "rp": 6,
    "key_vocabulary": 7,
    "ws_ms": 8,
    "ht_only": 9,
}

# Topic code to folder name mapping
TOPIC_FOLDERS = {
    "B1": "B1 - Cell Biology",
    "B2": "B2 - Organisation",
    "B3": "B3 - Infection and Response",
    "B4": "B4 - Bioenergetics",
    "B5": "B5 - Homeostasis and Response",
    "B6": "B6 - Inheritance Variation and Evolution",
    "B7": "B7 - Ecology",
    "C8": "C8 - Atomic Structure and Periodic Table",
    "C9": "C9 - Bonding Structure and Properties",
    "C10": "C10 - Quantitative Chemistry",
    "C11": "C11 - Chemical Changes",
    "C12": "C12 - Energy Changes",
    "C13": "C13 - Rate and Extent of Chemical Change",
    "C14": "C14 - Organic Chemistry",
    "C15": "C15 - Chemical Analysis",
    "C16": "C16 - Chemistry of the Atmosphere",
    "C17": "C17 - Using Resources",
    "P18": "P18 - Energy",
    "P19": "P19 - Electricity",
    "P20": "P20 - Particle Model of Matter",
    "P21": "P21 - Atomic Structure",
    "P22": "P22 - Forces",
    "P23": "P23 - Waves",
    "P24": "P24 - Magnetism and Electromagnetism",
}


def _cell_val(row, col_idx):
    """Get cell value, returning None for empty cells."""
    val = row[col_idx].value
    if val is None:
        return None
    val = str(val).strip()
    return val if val else None


def _extract_topic_code(topic_str):
    """Extract topic code like 'B1', 'C13', 'P22' from topic string."""
    if not topic_str:
        return None
    match = re.match(r"([BCP]\d+)", topic_str)
    return match.group(1) if match else None


def _spec_is_consolidation(spec_content):
    """Check if spec content describes consolidation rather than new teaching."""
    if not spec_content:
        return True
    return bool(CONSOLIDATION_SPEC_RE.match(spec_content.strip()))


def _is_booklet_lesson(title, spec_content, subject=None):
    """
    Determine if a row represents a booklet-worthy lesson.

    Strategy:
    - Hard patterns (assessment, DIRT, mock, etc.) always exclude
    - Soft patterns (review, consolidation) only exclude if spec content
      is empty or purely consolidation-focused
    - "All" subject lessons (contingency, exam readiness) are excluded
    """
    if not title:
        return False

    # Non-subject lessons
    if subject and subject.lower() == "all":
        return False

    # Hard exclude: always filter these out
    if HARD_NON_BOOKLET_RE.search(title):
        return False

    # Soft exclude: only filter if spec_content is consolidation/empty
    if SOFT_NON_BOOKLET_RE.search(title):
        if _spec_is_consolidation(spec_content):
            return False
        # Has real new content alongside review — keep as booklet lesson

    return True


def _get_subject_from_topic(topic_str):
    """Infer subject from topic code."""
    if not topic_str:
        return None
    code = _extract_topic_code(topic_str)
    if not code:
        return None
    if code.startswith("B"):
        return "Biology"
    elif code.startswith("C"):
        return "Chemistry"
    elif code.startswith("P"):
        return "Physics"
    return None


def parse_sheet(ws, year, header_row=4):
    """
    Parse a single worksheet (Year 10 or Year 11).

    Args:
        ws: openpyxl worksheet
        year: 10 or 11
        header_row: 1-indexed row number of the header

    Returns:
        List of lesson dicts
    """
    lessons = []
    current_week = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        # Skip rows that don't have a lesson number
        lesson_num_val = _cell_val(row, COL_MAP["lesson_num"])
        if not lesson_num_val:
            continue

        # Try to parse lesson number as int
        try:
            lesson_num = int(lesson_num_val)
        except (ValueError, TypeError):
            continue

        # Track current week
        week_val = _cell_val(row, COL_MAP["week"])
        if week_val:
            current_week = week_val

        title = _cell_val(row, COL_MAP["title"])
        spec_content = _cell_val(row, COL_MAP["spec_content"])
        subject = _cell_val(row, COL_MAP["subject"])
        topic = _cell_val(row, COL_MAP["topic"])
        rp = _cell_val(row, COL_MAP["rp"])
        key_vocab = _cell_val(row, COL_MAP["key_vocabulary"])
        ws_ms = _cell_val(row, COL_MAP["ws_ms"])
        ht_only = _cell_val(row, COL_MAP["ht_only"])

        # Handle subject from topic if the subject column references a review topic
        if subject and ":" not in (topic or "") and not _get_subject_from_topic(topic):
            # Subject might be in the subject column directly
            pass
        if not subject:
            subject = _get_subject_from_topic(topic)

        is_booklet = _is_booklet_lesson(title, spec_content, subject)

        # Extract topic code for folder mapping
        topic_code = _extract_topic_code(topic)
        topic_folder = TOPIC_FOLDERS.get(topic_code, "")

        # Build output folder path
        if subject and topic_folder:
            output_folder = f"{subject}/{topic_folder}/"
        else:
            output_folder = ""

        # Build filename
        if is_booklet and title:
            filename = f"L{lesson_num:03d} - {title}.docx"
        else:
            filename = ""

        lesson = {
            "year": year,
            "week": current_week,
            "lesson_number": lesson_num,
            "subject": subject,
            "topic": topic,
            "title": title,
            "spec_content": spec_content,
            "required_practical": rp,
            "key_vocabulary": key_vocab,
            "ws_ms": ws_ms,
            "ht_only": ht_only,
            "is_booklet_lesson": is_booklet,
            "filename": filename,
            "output_folder": output_folder,
            "prior_lessons": [],  # Populated after all lessons parsed
        }
        lessons.append(lesson)

    return lessons


def _populate_prior_lessons(lessons):
    """
    For each booklet lesson, populate prior_lessons with titles of
    all preceding booklet lessons in the same subject.
    """
    subject_history = {}  # subject -> list of titles

    for lesson in lessons:
        subject = lesson["subject"]
        if not subject:
            continue

        if subject not in subject_history:
            subject_history[subject] = []

        # Set prior lessons BEFORE adding current
        if lesson["is_booklet_lesson"]:
            lesson["prior_lessons"] = list(subject_history[subject])
            subject_history[subject].append(lesson["title"])


def parse_scheme_of_work(filepath):
    """
    Parse the full scheme of work spreadsheet.

    Args:
        filepath: Path to the .xlsx file

    Returns:
        dict with keys:
            'all_lessons': list of all lesson dicts (both years)
            'booklet_lessons': list of only booklet-worthy lessons
            'stats': summary statistics
    """
    filepath = Path(filepath)
    wb = load_workbook(filepath, read_only=True, data_only=True)

    all_lessons = []

    # Parse Year 10
    if "Year 10 Lessons" in wb.sheetnames:
        y10 = parse_sheet(wb["Year 10 Lessons"], year=10)
        all_lessons.extend(y10)

    # Parse Year 11
    if "Year 11 Lessons" in wb.sheetnames:
        y11 = parse_sheet(wb["Year 11 Lessons"], year=11)
        all_lessons.extend(y11)

    wb.close()

    # Populate prior lessons (across both years, in order)
    _populate_prior_lessons(all_lessons)

    # Filter booklet lessons
    booklet_lessons = [l for l in all_lessons if l["is_booklet_lesson"]]

    # Compute stats
    stats = {
        "total_rows": len(all_lessons),
        "booklet_lessons": len(booklet_lessons),
        "non_booklet_lessons": len(all_lessons) - len(booklet_lessons),
        "by_subject": {},
        "by_year": {10: 0, 11: 0},
    }

    for lesson in booklet_lessons:
        subj = lesson["subject"] or "Unknown"
        stats["by_subject"][subj] = stats["by_subject"].get(subj, 0) + 1
        stats["by_year"][lesson["year"]] = stats["by_year"].get(lesson["year"], 0) + 1

    return {
        "all_lessons": all_lessons,
        "booklet_lessons": booklet_lessons,
        "stats": stats,
    }


if __name__ == "__main__":
    import json
    import sys

    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else (
        "/Users/derek/Downloads/AQA_Combined_Science_Trilogy_Scheme_of_Work.xlsx"
    )

    result = parse_scheme_of_work(xlsx_path)

    print("=== PARSING STATS ===")
    print(json.dumps(result["stats"], indent=2))
    print()

    print("=== NON-BOOKLET LESSONS (filtered out) ===")
    for l in result["all_lessons"]:
        if not l["is_booklet_lesson"]:
            print(f"  Y{l['year']} L{l['lesson_number']:03d}: {l['title']}")
    print()

    print("=== FIRST 5 BOOKLET LESSONS ===")
    for l in result["booklet_lessons"][:5]:
        print(json.dumps(l, indent=2))
