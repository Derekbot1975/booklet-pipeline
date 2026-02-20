"""
Generic spreadsheet parser for any scheme of work.

Uses course config (from courses.py) to determine column mapping, sheet names,
topic folders, and subject inference. The existing AQA Combined Science logic
is preserved as one configuration among many.
"""

import re
from pathlib import Path
from openpyxl import load_workbook


# Hard non-booklet: these ALWAYS exclude regardless of spec content
HARD_NON_BOOKLET_PATTERNS = [
    r"\bassessment\b",
    r"\bdirt\b",
    r"\bmock\b",
    r"\brevision\b",
    r"\bexam\s+practice\b",
    r"\bwalking\s+talking\b",
    r"\bcontingency\b",
    r"\bexam\s+readiness\b",
]

# Soft non-booklet: only exclude if spec_content is missing or purely consolidation
SOFT_NON_BOOKLET_PATTERNS = [
    r"\breview\b",
    r"\bconsolidation\b",
    r"\brecap\b",
    r"\btopic\s+review\b",
    r"\bpractice\b",
]

HARD_NON_BOOKLET_RE = re.compile(
    "|".join(HARD_NON_BOOKLET_PATTERNS), re.IGNORECASE
)
SOFT_NON_BOOKLET_RE = re.compile(
    "|".join(SOFT_NON_BOOKLET_PATTERNS), re.IGNORECASE
)

# Spec content that indicates consolidation rather than new teaching
CONSOLIDATION_SPEC_RE = re.compile(
    r"^(consolidation|timed assessment|review|dedicated improvement|"
    r"full paper|teacher-led walkthrough|retrieval|rate calculations from graphs|"
    r"crude oil.*practice equations|all rps|review papers)",
    re.IGNORECASE,
)


def _cell_val(row, col_idx):
    """Get cell value, returning None for empty cells."""
    if col_idx is None or col_idx < 0 or col_idx >= len(row):
        return None
    val = row[col_idx].value
    if val is None:
        return None
    val = str(val).strip()
    return val if val else None


def _spec_is_consolidation(spec_content):
    """Check if spec content describes consolidation rather than new teaching."""
    if not spec_content:
        return True
    return bool(CONSOLIDATION_SPEC_RE.match(spec_content.strip()))


def _is_booklet_lesson(title, spec_content, subject=None):
    """Determine if a row represents a booklet-worthy lesson."""
    if not title:
        return False
    if subject and subject.lower() == "all":
        return False
    if HARD_NON_BOOKLET_RE.search(title):
        return False
    if SOFT_NON_BOOKLET_RE.search(title):
        if _spec_is_consolidation(spec_content):
            return False
    return True


def _extract_topic_code(topic_str, pattern):
    """Extract topic code from topic string using a regex pattern."""
    if not topic_str or not pattern:
        return None
    match = re.match(pattern, topic_str)
    return match.group(1) if match else None


def _get_subject_from_topic(topic_str, prefix_map, pattern):
    """Infer subject from topic code using prefix→subject mapping."""
    if not topic_str or not prefix_map or not pattern:
        return None
    code = _extract_topic_code(topic_str, pattern)
    if not code:
        return None
    # Try each prefix (longest first for specificity)
    for prefix in sorted(prefix_map.keys(), key=len, reverse=True):
        if code.startswith(prefix):
            return prefix_map[prefix]
    return None


def parse_sheet_generic(ws, year, course_config, header_row=None):
    """
    Parse a single worksheet using the course config for column mapping.

    Args:
        ws: openpyxl worksheet
        year: year group number (e.g. 7, 8, 9, 10, 11)
        course_config: dict with col_map, topic_folders, etc.
        header_row: override header row (1-indexed)

    Returns:
        List of lesson dicts
    """
    col_map = course_config.get("col_map", {})
    topic_folders = course_config.get("topic_folders", {})
    prefix_map = course_config.get("subject_from_topic_prefix", {})
    topic_pattern = course_config.get("topic_code_pattern", "")
    if header_row is None:
        header_row = course_config.get("header_row", 1)

    # Column indices (None if not configured)
    col_week = col_map.get("week")
    col_lesson_num = col_map.get("lesson_num")
    col_subject = col_map.get("subject")
    col_topic = col_map.get("topic")
    col_title = col_map.get("title")
    col_spec = col_map.get("spec_content")
    col_rp = col_map.get("rp")
    col_vocab = col_map.get("key_vocabulary")
    col_ws_ms = col_map.get("ws_ms")
    col_ht = col_map.get("ht_only")

    # The course may define a fixed subject (e.g. "Geography") for all rows
    fixed_subject = course_config.get("fixed_subject")

    lessons = []
    current_week = None

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1), start=header_row + 1
    ):
        # Must have a lesson number
        lesson_num_val = _cell_val(row, col_lesson_num)
        if not lesson_num_val:
            continue

        try:
            lesson_num = int(float(lesson_num_val))
        except (ValueError, TypeError):
            continue

        # Track week
        week_val = _cell_val(row, col_week)
        if week_val:
            current_week = week_val

        title = _cell_val(row, col_title)
        spec_content = _cell_val(row, col_spec)
        subject = _cell_val(row, col_subject)
        topic = _cell_val(row, col_topic)
        rp = _cell_val(row, col_rp)
        key_vocab = _cell_val(row, col_vocab)
        ws_ms = _cell_val(row, col_ws_ms)
        ht_only = _cell_val(row, col_ht)

        # Subject inference
        if fixed_subject:
            subject = fixed_subject
        elif not subject and prefix_map:
            subject = _get_subject_from_topic(topic, prefix_map, topic_pattern)

        is_booklet = _is_booklet_lesson(title, spec_content, subject)

        # Topic code → folder mapping
        topic_code = _extract_topic_code(topic, topic_pattern) if topic_pattern else None
        topic_folder = topic_folders.get(topic_code, "") if topic_code else ""

        # Output folder path
        if subject and topic_folder:
            output_folder = f"{subject}/{topic_folder}/"
        elif subject:
            output_folder = f"{subject}/"
        else:
            output_folder = ""

        # Filename
        if is_booklet and title:
            filename = f"L{lesson_num:03d} - {title}.docx"
        else:
            filename = ""

        lesson = {
            "year": year,
            "week": current_week,
            "lesson_number": lesson_num,
            "subject": subject,
            "topic": topic,
            "title": title,
            "spec_content": spec_content,
            "required_practical": rp,
            "key_vocabulary": key_vocab,
            "ws_ms": ws_ms,
            "ht_only": ht_only,
            "is_booklet_lesson": is_booklet,
            "filename": filename,
            "output_folder": output_folder,
            "prior_lessons": [],
        }
        lessons.append(lesson)

    return lessons


def _populate_prior_lessons(lessons):
    """
    For each booklet lesson, populate prior_lessons with titles of
    all preceding booklet lessons in the same subject.
    """
    subject_history = {}

    for lesson in lessons:
        subject = lesson["subject"]
        if not subject:
            continue

        if subject not in subject_history:
            subject_history[subject] = []

        if lesson["is_booklet_lesson"]:
            lesson["prior_lessons"] = list(subject_history[subject])
            subject_history[subject].append(lesson["title"])


def parse_course(course_config):
    """
    Parse a full scheme of work using the course config.

    Args:
        course_config: dict from courses.py

    Returns:
        dict with all_lessons, booklet_lessons, stats
    """
    filepath = Path(course_config["xlsx_path"])
    if not filepath.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {filepath}")

    wb = load_workbook(filepath, read_only=True, data_only=True)

    all_lessons = []
    sheets = course_config.get("sheets", [])

    for sheet_info in sheets:
        sheet_name = sheet_info["name"]
        year = sheet_info["year"]

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lessons = parse_sheet_generic(ws, year, course_config)
            all_lessons.extend(lessons)

    wb.close()

    # Populate prior lessons
    _populate_prior_lessons(all_lessons)

    # Filter booklet lessons
    booklet_lessons = [l for l in all_lessons if l["is_booklet_lesson"]]

    # Compute stats
    stats = {
        "total_rows": len(all_lessons),
        "booklet_lessons": len(booklet_lessons),
        "non_booklet_lessons": len(all_lessons) - len(booklet_lessons),
        "by_subject": {},
        "by_year": {},
    }

    for lesson in booklet_lessons:
        subj = lesson["subject"] or "Unknown"
        stats["by_subject"][subj] = stats["by_subject"].get(subj, 0) + 1
        yr = lesson["year"]
        stats["by_year"][yr] = stats["by_year"].get(yr, 0) + 1

    return {
        "all_lessons": all_lessons,
        "booklet_lessons": booklet_lessons,
        "stats": stats,
    }


def preview_spreadsheet(xlsx_path, sheet_name=None, max_rows=10):
    """
    Preview a spreadsheet to help the user configure column mapping.

    Returns:
        dict with sheet_names, headers, sample_rows
    """
    filepath = Path(xlsx_path)
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = load_workbook(filepath, read_only=True, data_only=True)

    result = {
        "sheet_names": wb.sheetnames,
        "sheets": {},
    }

    target_sheets = [sheet_name] if sheet_name else wb.sheetnames

    for sn in target_sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows_data = []
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows + 5)):
            cells = []
            for cell in row:
                val = cell.value
                cells.append(str(val).strip() if val is not None else "")
            rows_data.append(cells)

        result["sheets"][sn] = rows_data

    wb.close()
    return result
