"""
Scheme of Work Generator & Review Engine (Prompt Sheet 13).

Provides:
  - AI review of uploaded schemes against reference documents
  - AI generation of new schemes of work from scratch
  - Structured JSON storage of schemes
  - Apply-suggestion support for iterative improvement
"""

import json
import logging
import os
import re
import time
import uuid
from datetime import datetime
from pathlib import Path

import anthropic
from dotenv import load_dotenv

load_dotenv(override=True)

logger = logging.getLogger(__name__)


def _extract_json(raw, expect_object=False):
    """Robustly extract JSON from AI response text."""
    # Strip all markdown code fences (may appear anywhere, not just start/end)
    raw = re.sub(r"```(?:json)?\s*\n?", "", raw)
    raw = raw.strip()
    if expect_object:
        start, end = raw.find("{"), raw.rfind("}")
    else:
        start, end = raw.find("["), raw.rfind("]")
    if start != -1 and end != -1 and end > start:
        raw = raw[start:end + 1]
    elif start != -1:
        # Truncated output — take from first bracket to end
        raw = raw[start:]
    raw = re.sub(r",\s*([}\]])", r"\1", raw)
    return raw


def _repair_json(raw_text, expect_object=True):
    """Attempt to repair malformed JSON from AI responses.

    Handles:
    - Truncated output (missing closing brackets/braces)
    - Missing commas between elements
    - Single quotes instead of double quotes
    - Unquoted keys
    - Trailing commas (already handled by _extract_json)
    """
    text = _extract_json(raw_text, expect_object=expect_object)

    # First try: maybe it's already valid
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Fix: missing commas between } { or ] [ or } [ or ] {
    text = re.sub(r'(\})\s*(\{)', r'\1,\2', text)
    text = re.sub(r'(\])\s*(\[)', r'\1,\2', text)
    text = re.sub(r'(\})\s*(\[)', r'\1,\2', text)
    text = re.sub(r'(\])\s*(\{)', r'\1,\2', text)

    # Fix: missing commas between "value" "key" patterns (string followed by string)
    text = re.sub(r'"\s*\n\s*"', '",\n"', text)

    # Fix: missing comma after true/false/null/number before "key"
    text = re.sub(r'(true|false|null|\d)\s*\n\s*"', r'\1,\n"', text)

    # Second try after comma fixes
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Fix truncated JSON: count brackets and close any unclosed ones
    open_braces = text.count('{') - text.count('}')
    open_brackets = text.count('[') - text.count(']')

    if open_braces > 0 or open_brackets > 0:
        # Truncated output — try progressively aggressive repairs

        # Strategy 1: Trim trailing partial content, then close brackets
        # using stack-based nesting order
        for trim_text in [text, text.rstrip().rstrip(',')]:
            # Try trimming back to the last complete value
            for trim_pattern in [
                r',\s*"[^"]*$',             # trailing partial key/string
                r',\s*"[^"]*":\s*"[^"]*$',  # trailing partial key:"value
                r',\s*"[^"]*":\s*\{[^}]*$', # trailing partial key:{obj
                r',\s*"[^"]*":\s*\[[^\]]*$',# trailing partial key:[arr
                r',\s*\{[^}]*$',            # trailing partial object in array
                r'',                         # no trim (try as-is)
            ]:
                if trim_pattern:
                    candidate = re.sub(trim_pattern, '', trim_text)
                else:
                    candidate = trim_text

                # Use stack to determine correct bracket close order
                stack = []
                in_string = False
                escape_next = False
                for ch in candidate:
                    if escape_next:
                        escape_next = False
                        continue
                    if ch == '\\' and in_string:
                        escape_next = True
                        continue
                    if ch == '"':
                        in_string = not in_string
                        continue
                    if in_string:
                        continue
                    if ch in '{[':
                        stack.append('}' if ch == '{' else ']')
                    elif ch in '}]':
                        if stack:
                            stack.pop()

                if stack:
                    candidate += ''.join(reversed(stack))

                try:
                    return json.loads(candidate)
                except json.JSONDecodeError:
                    continue

    # If nothing worked, raise the original error for logging
    raise json.JSONDecodeError(
        f"Could not repair JSON ({open_braces} unclosed braces, {open_brackets} unclosed brackets)",
        text[:200], 0
    )


SCHEMES_DIR = Path(__file__).parent / "data" / "schemes"
SCHEMES_DIR.mkdir(parents=True, exist_ok=True)


# ───────────────────────────────────────────────────────────────────
# System prompts
# ───────────────────────────────────────────────────────────────────

SOW_REVIEW_SYSTEM_PROMPT = """You are a UK curriculum expert and school improvement specialist.
You review schemes of work against reference materials and provide detailed, constructive feedback.

RULES:
- Use UK English throughout.
- Be specific and actionable in every piece of feedback.
- Ground every suggestion in the reference materials provided.
- Be constructive — highlight strengths as well as improvements.
- When suggesting changes, be precise about where they should go.

OUTPUT: You MUST return valid JSON matching the schema described in the user prompt.
Do NOT wrap the JSON in markdown code fences. Return raw JSON only."""

SOW_GENERATE_SYSTEM_PROMPT = """You are a UK curriculum design expert who creates exceptionally detailed, research-informed schemes of work at the quality level expected by MATs (Multi Academy Trusts) and Ofsted deep dives.

You must produce output matching the depth and rigour of the following exemplar structure (adapted for any subject):

MULTI-SHEET STRUCTURE (you generate all data for these sheets):

1. OVERVIEW — Includes:
   - Lesson structure (e.g. retrieval starter, core teaching, exit ticket timings)
   - Design rationale with named research/curriculum thinkers relevant to the subject
   - Content reduction decisions (what was cut and why, citing "less is more" principles)
   - Sequencing rationale (5-10 numbered principles with researcher citations)
   - Total lessons, assessment model, consolidation model

2. LESSONS (one set per year group) — Every lesson MUST have ALL of these columns:
   - Week number (Wk)
   - Lesson number (L#) — sequential across the whole year
   - Enquiry / Unit — the overarching enquiry question or unit name
   - Lesson Title — specific, descriptive title
   - Specification Content & Learning Objectives — rich prose description (2-4 sentences) of what is taught, NOT just bullet points
   - Key Vocabulary (Pre-taught) — 4-8 Tier 2 and Tier 3 terms per lesson, comma-separated
   - Subject-Specific Skill / Concept — the disciplinary or procedural skill being developed (e.g. for History: "Causation", for Science: "Working Scientifically", for Maths: "Reasoning", for English: "Analytical Writing")
   - Progression Level — explicit progression descriptor (e.g. "L1: Identifying a cause", "L3: Categorised comparison")
   - Substantive Concepts — the big ideas / concept threads running through the curriculum (e.g. Power, Energy, Number, Genre)
   - Retrieval Starter Focus — what prior knowledge the low-stakes quiz covers

   SPECIAL LESSON TYPES (distribute appropriately):
   - Assessment lessons: 2-3 per year, marked "A" in lesson type
   - DIRT / Feedback lessons: after each assessment, marked "A"
   - Writing/Skill Instruction lessons: explicit teaching of the skill BEFORE assessment, marked "W"
   - Consolidation weeks: 1 per half-term (~6 per year), marked "C", for pure retrieval, review, extended practice

3. PROGRESSION LADDERS — Subject-specific skill/concept progression with 5 levels:
   - Each level: Descriptor, Worked Example from the scheme, Common Misconception, Pupil-Friendly Version
   - For History: Causation, Change & Continuity, Significance, Evidence, Interpretation, Similarity & Difference
   - For Science: Working Scientifically, Mathematical Skills, Practical Skills, Scientific Vocabulary
   - For Maths: Fluency, Reasoning, Problem Solving, Mathematical Communication
   - For English: Reading Analysis, Writing Composition, Spelling/Punctuation/Grammar, Spoken Language
   - For other subjects: identify 4-6 subject-appropriate disciplinary skills/concepts

4. CURRICULUM MAP — Substantive concept threads tracked across all year groups showing how big ideas develop

5. VOCABULARY TEACHING — Structured approach to:
   - Pre-teaching routine (steps, timing)
   - Cumulative vocabulary in retrieval starters (question types, frequency)
   - Deliberate re-use plan (key terms, where introduced, where re-used, why)

CRITICAL RULES:
- Use UK English throughout.
- Cover ALL required National Curriculum or specification content for the year group.
- Sequence lessons logically with research-informed progression (cite researchers by name).
- Every lesson must have ALL columns filled — no empty fields.
- Lessons per year must exactly match the total specified.
- Include retrieval practice in EVERY lesson (cumulative, interleaved across prior units).
- Include consolidation weeks (1 per half-term).
- Include explicit writing/skill instruction lessons BEFORE assessments.
- Vocabulary must be Tier 2 AND Tier 3, pre-taught and cumulatively revisited.
- Substantive concept threads must be traceable across the whole year.
- Progression levels must increase across the year (most pupils L2-3 by end of year 7, L3-4 by year 8, L4-5 by year 9 or equivalent for the key stage).
- Reference the school's quality framework documents if provided.
- If a specification or NC document is provided, ensure COMPLETE coverage of required content.

OUTPUT: You MUST return valid JSON matching the schema described in the user prompt.
Do NOT wrap the JSON in markdown code fences. Return raw JSON only."""

SOW_APPLY_SUGGESTION_PROMPT = """You are a UK curriculum expert modifying a scheme of work based on a specific suggestion.

RULES:
- Make ONLY the change described in the suggestion.
- Preserve all other content exactly as it is.
- Maintain consistent formatting and structure.
- If adding lessons, number them appropriately.
- If moving content, update all affected lesson numbers.
- Use UK English throughout.

OUTPUT: Return the COMPLETE modified scheme of work as valid JSON.
Do NOT wrap the JSON in markdown code fences. Return raw JSON only."""


def _get_client():
    """Get Anthropic client."""
    return anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))


def _create_message(**kwargs):
    """Create a message using streaming to avoid the 10-minute SDK timeout."""
    from ai_client import create_message
    return create_message(**kwargs)


def _scheme_path(scheme_id):
    return SCHEMES_DIR / f"{scheme_id}.json"


def _review_path(scheme_id):
    return SCHEMES_DIR / f"{scheme_id}_review.json"


# ───────────────────────────────────────────────────────────────────
# Scheme CRUD
# ───────────────────────────────────────────────────────────────────

def list_schemes(deduplicate=True):
    """List all saved schemes. When deduplicate=True, keeps only the latest
    version of each unique (title + subject + yearGroup) combination."""
    raw = []
    for p in sorted(SCHEMES_DIR.glob("*.json")):
        if p.stem.endswith("_review"):
            continue
        try:
            data = json.loads(p.read_text())
            raw.append({
                "id": data.get("id", p.stem),
                "subject": data.get("subject", ""),
                "key_stage": data.get("keyStage", data.get("key_stage", "")),
                "year_group": data.get("yearGroup", data.get("year_group", "")),
                "title": data.get("title", f"{data.get('subject', 'Unknown')} — Year {data.get('yearGroup', '?')}"),
                "total_lessons": data.get("totalLessons", 0),
                "exam_board": data.get("examBoard", data.get("exam_board", "")),
                "created_at": data.get("created_at", ""),
                "updated_at": data.get("updated_at", ""),
                "source": data.get("source", ""),
                "has_review": _review_path(data.get("id", p.stem)).exists(),
                "file_path": str(p),
            })
        except (json.JSONDecodeError, KeyError):
            continue

    if not deduplicate:
        return raw

    # Deduplicate: keep only the latest (by updated_at or created_at) per
    # unique (title, subject, year_group) key
    best = {}
    for s in raw:
        key = (s["title"].strip().lower(), s["subject"].strip().lower(),
               str(s["year_group"]).strip())
        ts = s.get("updated_at") or s.get("created_at") or ""
        if key not in best or ts > (best[key].get("updated_at") or best[key].get("created_at") or ""):
            best[key] = s
    return list(best.values())


def get_scheme(scheme_id):
    """Load a full scheme by ID."""
    path = _scheme_path(scheme_id)
    if not path.exists():
        return None
    return json.loads(path.read_text())


def save_scheme(scheme_data):
    """Save a scheme (creates or updates)."""
    if "id" not in scheme_data or not scheme_data["id"]:
        scheme_data["id"] = str(uuid.uuid4())[:8]
    now = datetime.utcnow().isoformat()
    if "created_at" not in scheme_data:
        scheme_data["created_at"] = now
    scheme_data["updated_at"] = now
    _scheme_path(scheme_data["id"]).write_text(json.dumps(scheme_data, indent=2))
    return scheme_data


def delete_scheme(scheme_id):
    """Delete a scheme and its review."""
    path = _scheme_path(scheme_id)
    if path.exists():
        path.unlink()
    review = _review_path(scheme_id)
    if review.exists():
        review.unlink()
    return True


def get_review(scheme_id):
    """Load review results for a scheme."""
    path = _review_path(scheme_id)
    if not path.exists():
        return None
    return json.loads(path.read_text())


# ───────────────────────────────────────────────────────────────────
# Import from spreadsheet
# ───────────────────────────────────────────────────────────────────

def find_existing_scheme(title, subject, year_group):
    """Check if a scheme with the same title+subject+yearGroup already exists.
    Returns the existing scheme data (dict) or None."""
    title_lc = (title or "").strip().lower()
    subject_lc = (subject or "").strip().lower()
    yg_str = str(year_group).strip()
    for p in SCHEMES_DIR.glob("*.json"):
        if p.stem.endswith("_review"):
            continue
        try:
            data = json.loads(p.read_text())
            if (data.get("title", "").strip().lower() == title_lc
                    and data.get("subject", "").strip().lower() == subject_lc
                    and str(data.get("yearGroup", data.get("year_group", ""))).strip() == yg_str):
                return data
        except (json.JSONDecodeError, KeyError):
            continue
    return None


def import_from_course(course_config, parsed_data):
    """
    Convert existing course parsed data into a scheme of work JSON structure.
    This bridges the existing pipeline data to the scheme format.

    Deduplication: if a scheme with the same title+subject+yearGroup already
    exists, update it in-place rather than creating a new copy.
    """
    lessons = parsed_data.get("all_lessons", [])
    if not lessons:
        raise ValueError("No lessons found in parsed data")

    # Group by year, then by topic/unit
    by_year = {}
    for l in lessons:
        y = l.get("year", 0)
        by_year.setdefault(y, []).append(l)

    # Build terms/units structure
    terms = []
    for year in sorted(by_year.keys()):
        year_lessons = sorted(by_year[year], key=lambda x: x.get("lesson_number", 0))

        # Group by topic
        units = []
        current_topic = None
        current_unit_lessons = []

        for l in year_lessons:
            topic = l.get("topic", "General")
            if topic != current_topic:
                if current_unit_lessons:
                    units.append({
                        "title": current_topic or "General",
                        "lessons": current_unit_lessons,
                        "endOfUnitAssessment": False,
                    })
                current_topic = topic
                current_unit_lessons = []

            current_unit_lessons.append({
                "number": l.get("lesson_number", 0),
                "title": l.get("title", ""),
                "objectives": [l.get("spec_content", "")] if l.get("spec_content") else [],
                "keyVocabulary": [v.strip() for v in (l.get("key_vocabulary") or "").split(",") if v.strip()],
                "resources": "",
                "assessment": None,
                "crossCurricular": [],
            })

        if current_unit_lessons:
            units.append({
                "title": current_topic or "General",
                "lessons": current_unit_lessons,
                "endOfUnitAssessment": False,
            })

        terms.append({
            "name": f"Year {year}",
            "units": units,
        })

    title = course_config.get("name", "Imported Scheme")
    subject = course_config.get("subjects", [""])[0] if course_config.get("subjects") else ""
    year_group = sorted(by_year.keys())[0] if by_year else 0

    # Check for existing scheme with same title+subject+yearGroup
    existing = find_existing_scheme(title, subject, year_group)

    scheme = {
        "id": existing["id"] if existing else str(uuid.uuid4())[:8],
        "subject": subject,
        "keyStage": f"KS{course_config.get('key_stage', '')}" if course_config.get("key_stage") else "",
        "yearGroup": year_group,
        "examBoard": course_config.get("exam_board", ""),
        "totalLessons": len(lessons),
        "title": title,
        "source": "imported",
        "terms": terms,
        "created_at": existing.get("created_at", datetime.utcnow().isoformat()) if existing else datetime.utcnow().isoformat(),
    }

    saved = save_scheme(scheme)
    if existing:
        logger.info(f"Updated existing scheme {existing['id']} instead of creating duplicate")
    return saved


def cleanup_duplicate_schemes():
    """One-time cleanup: find duplicate schemes (same title+subject+yearGroup)
    and keep only the latest, deleting the rest. Returns count of deleted."""
    groups = {}
    for p in sorted(SCHEMES_DIR.glob("*.json")):
        if p.stem.endswith("_review"):
            continue
        try:
            data = json.loads(p.read_text())
            key = (
                data.get("title", "").strip().lower(),
                data.get("subject", "").strip().lower(),
                str(data.get("yearGroup", data.get("year_group", ""))).strip(),
            )
            ts = data.get("updated_at") or data.get("created_at") or ""
            groups.setdefault(key, []).append((ts, p, data.get("id", p.stem)))
        except (json.JSONDecodeError, KeyError):
            continue

    deleted = 0
    for key, entries in groups.items():
        if len(entries) <= 1:
            continue
        # Sort by timestamp desc — keep the newest
        entries.sort(key=lambda x: x[0], reverse=True)
        for ts, p, sid in entries[1:]:
            logger.info(f"Deleting duplicate scheme {sid} ({p.name})")
            p.unlink(missing_ok=True)
            # Also remove its review if present
            review = _review_path(sid)
            if review.exists():
                review.unlink()
            deleted += 1
    return deleted


def import_from_file(file_path, subject=None, key_stage=None, year_group=None,
                     exam_board=None):
    """Import a scheme from an uploaded file (xlsx, docx, csv).
    Parses the file into the scheme JSON structure."""
    ext = Path(file_path).suffix.lower()

    lessons = []
    if ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for ws in wb.worksheets:
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
                title = (row_dict.get("title") or row_dict.get("lesson title") or
                         row_dict.get("lesson") or "")
                if not title:
                    continue
                lessons.append({
                    "lesson_number": len(lessons) + 1,
                    "title": str(title),
                    "topic": str(row_dict.get("topic") or row_dict.get("unit") or "General"),
                    "spec_content": str(row_dict.get("spec content") or row_dict.get("objectives") or
                                       row_dict.get("content") or ""),
                    "key_vocabulary": str(row_dict.get("key vocabulary") or row_dict.get("vocabulary") or ""),
                    "year": int(year_group) if year_group else 0,
                })
    elif ext == ".csv":
        import csv
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_lc = {k.strip().lower(): v for k, v in row.items()}
                title = (row_lc.get("title") or row_lc.get("lesson title") or
                         row_lc.get("lesson") or "")
                if not title:
                    continue
                lessons.append({
                    "lesson_number": len(lessons) + 1,
                    "title": str(title),
                    "topic": str(row_lc.get("topic") or row_lc.get("unit") or "General"),
                    "spec_content": str(row_lc.get("spec content") or row_lc.get("objectives") or ""),
                    "key_vocabulary": str(row_lc.get("key vocabulary") or row_lc.get("vocabulary") or ""),
                    "year": int(year_group) if year_group else 0,
                })
    else:
        raise ValueError(f"Unsupported file type: {ext}. Please upload .xlsx, .xls, or .csv")

    if not lessons:
        raise ValueError("No lessons found in uploaded file. Check column headers include 'Title' or 'Lesson'.")

    # Build terms/units structure
    units = []
    current_topic = None
    current_lessons = []
    for l in lessons:
        topic = l.get("topic", "General")
        if topic != current_topic:
            if current_lessons:
                units.append({"title": current_topic or "General",
                              "lessons": current_lessons,
                              "endOfUnitAssessment": False})
            current_topic = topic
            current_lessons = []
        current_lessons.append({
            "number": l["lesson_number"],
            "title": l["title"],
            "objectives": [l["spec_content"]] if l["spec_content"] else [],
            "keyVocabulary": [v.strip() for v in l["key_vocabulary"].split(",") if v.strip()],
            "resources": "",
            "assessment": None,
            "crossCurricular": [],
        })
    if current_lessons:
        units.append({"title": current_topic or "General",
                      "lessons": current_lessons,
                      "endOfUnitAssessment": False})

    title_str = f"{subject or 'Imported'} — Year {year_group or '?'} ({key_stage or ''})"

    # Check for existing duplicate
    existing = find_existing_scheme(title_str, subject, year_group)

    scheme = {
        "id": existing["id"] if existing else str(uuid.uuid4())[:8],
        "subject": subject or "",
        "keyStage": key_stage or "",
        "yearGroup": int(year_group) if year_group else 0,
        "examBoard": exam_board or "",
        "totalLessons": len(lessons),
        "title": title_str,
        "source": "file_upload",
        "terms": [{"name": f"Year {year_group or '?'}", "units": units}],
        "created_at": existing.get("created_at", datetime.utcnow().isoformat()) if existing else datetime.utcnow().isoformat(),
    }

    return save_scheme(scheme)


# ───────────────────────────────────────────────────────────────────
# AI Review
# ───────────────────────────────────────────────────────────────────

def review_scheme(scheme_id, reference_context="", model="claude-sonnet-4-5-20250929"):
    """
    Run an AI review of a scheme of work against reference documents.

    Returns structured review with strengths, improvements, missing content,
    and suggested resequencing.
    """
    scheme = get_scheme(scheme_id)
    if not scheme:
        raise ValueError(f"Scheme not found: {scheme_id}")

    start = time.time()

    user_prompt = f"""Review this scheme of work against the reference materials provided.

REFERENCE MATERIALS:
{reference_context if reference_context else "(No reference documents uploaded yet. Review based on general UK curriculum best practice.)"}

SCHEME OF WORK TO REVIEW:
{json.dumps(scheme, indent=2)}

REVIEW AGAINST THESE CRITERIA:
1. CURRICULUM COVERAGE — Does it cover all required content? Any gaps?
2. SEQUENCING & PROGRESSION — Is the lesson sequence logical? Does prior knowledge build?
3. EXPERT INPUT ALIGNMENT — Does it follow subject-specific pedagogical guidance?
4. QUALITY FRAMEWORK ALIGNMENT — Does it meet quality criteria?
5. BALANCE & BREADTH — Good mix of knowledge and skills? Assessment points?

Return your review as JSON with this EXACT structure:
{{
    "overallRating": "excellent" or "good" or "requires_improvement" or "inadequate",
    "overallSummary": "2-3 sentence summary",
    "strengths": [
        {{
            "area": "string",
            "detail": "specific praise",
            "evidence": "which part of the SoW demonstrates this"
        }}
    ],
    "improvements": [
        {{
            "area": "string",
            "issue": "what is wrong or missing",
            "suggestion": "specific actionable improvement",
            "priority": "critical" or "important" or "nice_to_have",
            "reference": "which reference document this comes from (or general best practice)",
            "affectedLessons": [1, 2, 3] or null
        }}
    ],
    "missingContent": [
        {{
            "topic": "string",
            "source": "National Curriculum / Specification / best practice",
            "suggestedPlacement": "After Unit X, Lesson Y",
            "estimatedLessons": 2
        }}
    ],
    "suggestedResequencing": [
        {{
            "current": "description of current placement",
            "suggested": "suggested new placement",
            "rationale": "why this change would improve the scheme"
        }}
    ]
}}"""

    messages = [{"role": "user", "content": user_prompt}]

    system_parts = [
        {
            "type": "text",
            "text": SOW_REVIEW_SYSTEM_PROMPT,
            "cache_control": {"type": "ephemeral"},
        }
    ]

    response = _create_message(
        model=model,
        max_tokens=8000,
        system=system_parts,
        messages=messages,
    )

    duration = round(time.time() - start, 1)
    raw_text = response.content[0].text.strip()

    try:
        review = _repair_json(raw_text, expect_object=True)
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse review JSON: {e}\nRaw: {raw_text[:800]}")
        review = {
            "overallRating": "requires_improvement",
            "overallSummary": "Review generated but output could not be parsed as JSON.",
            "rawResponse": raw_text,
            "strengths": [],
            "improvements": [],
            "missingContent": [],
            "suggestedResequencing": [],
        }

    # Save review
    review["reviewed_at"] = datetime.utcnow().isoformat()
    review["model"] = model
    review["duration_s"] = duration
    review["usage"] = {
        "input_tokens": response.usage.input_tokens,
        "output_tokens": response.usage.output_tokens,
    }

    _review_path(scheme_id).write_text(json.dumps(review, indent=2))
    return review


# ───────────────────────────────────────────────────────────────────
# AI Generation
# ───────────────────────────────────────────────────────────────────

def generate_scheme(subject, key_stage, year_group, lessons_per_week=3,
                    weeks_per_term=6, exam_board=None, priorities=None,
                    exclusions=None, reference_context="",
                    model="claude-sonnet-4-5-20250929"):
    """
    Generate a complete scheme of work from scratch using AI.

    Produces exemplar-quality output with multi-sheet structure:
    overview, lessons, progression ladders, curriculum map, vocabulary guidance.

    Returns saved scheme data.
    """
    start = time.time()

    total_terms = 6  # UK academic year: 6 half-terms
    total_lessons = lessons_per_week * weeks_per_term * total_terms
    term_names = ["Autumn 1", "Autumn 2", "Spring 1", "Spring 2", "Summer 1", "Summer 2"]

    user_prompt = f"""Generate a complete, exemplar-quality scheme of work for:
- Subject: {subject}
- Key Stage: {key_stage}
- Year Group: {year_group}
- Exam Board: {exam_board or "N/A"}
- Lessons per week: {lessons_per_week}
- Weeks per term (half-term): {weeks_per_term}
- Total lessons across the year: {total_lessons}
- Terms: {', '.join(term_names)}
- Weeks per year: {total_lessons // lessons_per_week}

{f"PRIORITISE THESE TOPICS: {priorities}" if priorities else ""}
{f"EXCLUDE THESE TOPICS: {exclusions}" if exclusions else ""}

REFERENCE MATERIALS:
{reference_context if reference_context else "(No reference documents uploaded. Generate based on UK National Curriculum / specification best practice.)"}

Return the scheme as JSON with this EXACT structure (all fields mandatory):
{{
    "subject": "{subject}",
    "keyStage": "{key_stage}",
    "yearGroup": {year_group},
    "examBoard": {json.dumps(exam_board)},
    "totalLessons": {total_lessons},

    "overview": {{
        "title": "{subject} {key_stage} Year {year_group} Scheme of Work",
        "subtitle": "Year {year_group} — [period/theme description] | {total_lessons} lessons",
        "lessonStructure": [
            "Minutes 0-10: CUMULATIVE RETRIEVAL (low-stakes, interleaved across all prior units)",
            "Minutes 10-50: CORE TEACHING (enquiry-driven, booklet-supported)",
            "Minutes 50-55: EXIT TICKET / KNOWLEDGE CHECK"
        ],
        "designRationale": [
            "1. [Principle] — cite relevant researcher/thinker by name",
            "2. [Principle] — cite relevant researcher/thinker by name"
        ],
        "contentDecisions": [
            "[What was included/excluded and why, citing 'less is more' principles]"
        ],
        "assessmentModel": "Description of assessment approach (e.g. 2 per year + formative checkpoints)",
        "consolidationModel": "Description (e.g. 1 consolidation week per half-term)"
    }},

    "units": [
        {{
            "name": "Autumn 1",
            "enquiryTitle": "Unit enquiry question or theme title",
            "lessons": [
                {{
                    "week": 1,
                    "number": 1,
                    "unit": "Enquiry / unit title",
                    "title": "Specific lesson title",
                    "content": "Rich 2-4 sentence description of specification content and learning objectives",
                    "vocabulary": "term1, term2, term3, term4, term5, term6",
                    "skill": "Subject-specific skill (e.g. Causation, Working Scientifically, Reasoning)",
                    "progressionLevel": "L2: Description of what this level looks like",
                    "concepts": "Big idea thread 1; Big idea thread 2",
                    "retrievalFocus": "R: What prior knowledge the retrieval quiz covers",
                    "lessonType": null
                }}
            ]
        }}
    ],

    "progressionLadders": [
        {{
            "skill": "Name of disciplinary skill/concept",
            "yearExpectations": "Most Year {year_group} pupils → L[x]-[y] by end of year",
            "levels": [
                {{
                    "level": "L1",
                    "name": "Short name",
                    "descriptor": "What pupils can do at this level",
                    "workedExample": "Concrete example from THIS scheme showing L1 work",
                    "misconception": "What pupils commonly get wrong at this level",
                    "pupilFriendly": "I can..."
                }},
                {{
                    "level": "L2",
                    "name": "Short name",
                    "descriptor": "...",
                    "workedExample": "...",
                    "misconception": "...",
                    "pupilFriendly": "I can..."
                }},
                {{
                    "level": "L3",
                    "name": "Short name",
                    "descriptor": "...",
                    "workedExample": "...",
                    "misconception": "...",
                    "pupilFriendly": "I can..."
                }},
                {{
                    "level": "L4",
                    "name": "Short name",
                    "descriptor": "...",
                    "workedExample": "...",
                    "misconception": "...",
                    "pupilFriendly": "I can..."
                }},
                {{
                    "level": "L5",
                    "name": "Short name",
                    "descriptor": "...",
                    "workedExample": "...",
                    "misconception": "...",
                    "pupilFriendly": "I can..."
                }}
            ]
        }}
    ],

    "curriculumMap": [
        {{
            "thread": "Substantive concept thread name (e.g. Power & Authority, Energy, Genre)",
            "coverage": "Brief description of how this thread develops across the year"
        }}
    ],

    "vocabularyTeaching": {{
        "preTeachingRoutine": [
            {{"step": "1. Display", "teacher": "Show 6-8 key terms", "pupil": "Read terms", "time": "30 sec"}},
            {{"step": "2. Define", "teacher": "Give clear definition", "pupil": "Write in glossary", "time": "2 min"}},
            {{"step": "3. Contextualise", "teacher": "Use in subject sentence", "pupil": "Hear in context", "time": "1 min"}},
            {{"step": "4. Practice", "teacher": "Ask for sentence using term", "pupil": "Write and share", "time": "1.5 min"}}
        ],
        "retrievalQuestionTypes": [
            {{"type": "Define", "example": "Define '[term]'. Use it in a sentence about [topic].", "frequency": "Every lesson"}},
            {{"type": "Match", "example": "Match terms to definitions (mix current + prior units)", "frequency": "Weekly"}},
            {{"type": "Use in context", "example": "Explain [topic] using these words: [term1], [term2], [term3]", "frequency": "Fortnightly"}},
            {{"type": "Odd one out", "example": "Which is the odd one out: [term1], [term2], [term3]? Explain.", "frequency": "Monthly"}}
        ],
        "deliberateReuse": [
            {{"term": "Key term", "introduced": "When/where", "reusedIn": "Where it reappears", "why": "Why this matters"}}
        ]
    }}
}}

CRITICAL REQUIREMENTS:
1. Generate EXACTLY {total_lessons} lessons numbered 1 to {total_lessons}.
2. Every lesson MUST have ALL fields filled (content, vocabulary, skill, progressionLevel, concepts, retrievalFocus).
3. The "content" field must be 2-4 rich sentences, NOT just a title or bullet points.
4. Include 4-6 progression ladders appropriate for {subject}.
5. Include 5-8 substantive concept threads in the curriculum map.
6. Distribute lesson types: ~{total_lessons - 12 - 6} core lessons, ~6 assessment+DIRT pairs (A), ~6 consolidation (C), ~4 writing/skill instruction (W).
7. Lessons labelled "lessonType": "A" for assessment/DIRT, "C" for consolidation, "W" for explicit skill instruction, null for core teaching.
8. Retrieval starters must reference specific prior content, becoming more interleaved as the year progresses.
9. Progression levels must increase across the year.
10. Vocabulary must include both Tier 2 (academic) and Tier 3 (subject-specific) terms."""

    system_parts = [
        {
            "type": "text",
            "text": SOW_GENERATE_SYSTEM_PROMPT,
            "cache_control": {"type": "ephemeral"},
        }
    ]

    response = _create_message(
        model=model,
        max_tokens=64000,
        system=system_parts,
        messages=[{"role": "user", "content": user_prompt}],
    )

    duration = round(time.time() - start, 1)
    raw_text = response.content[0].text.strip()
    stop_reason = response.stop_reason

    if stop_reason == "max_tokens":
        logger.warning(f"Scheme generation hit max_tokens — output likely truncated ({len(raw_text)} chars)")

    try:
        scheme_data = _repair_json(raw_text, expect_object=True)
        if stop_reason == "max_tokens":
            logger.info("Successfully repaired truncated JSON from scheme generation")
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse generated SoW JSON: {e}\nRaw: {raw_text[:800]}")
        raise ValueError(f"AI generated invalid JSON for scheme: {e}")

    # Backward-compat: build legacy "terms" structure from new "units" if present
    if "units" in scheme_data and "terms" not in scheme_data:
        scheme_data["terms"] = []
        for unit_block in scheme_data["units"]:
            term_entry = {
                "name": unit_block.get("name", ""),
                "units": [{
                    "title": unit_block.get("enquiryTitle", unit_block.get("name", "")),
                    "lessons": [],
                    "endOfUnitAssessment": any(
                        l.get("lessonType") == "A" for l in unit_block.get("lessons", [])
                    ),
                }],
            }
            for lesson in unit_block.get("lessons", []):
                term_entry["units"][0]["lessons"].append({
                    "number": lesson.get("number", 0),
                    "title": lesson.get("title", ""),
                    "objectives": [lesson.get("content", "")],
                    "keyVocabulary": [v.strip() for v in lesson.get("vocabulary", "").split(",") if v.strip()],
                    "resources": "",
                    "assessment": lesson.get("lessonType") if lesson.get("lessonType") == "A" else None,
                    "crossCurricular": [],
                })
            scheme_data["terms"].append(term_entry)

    # Add metadata
    scheme_data["id"] = str(uuid.uuid4())[:8]
    scheme_data["source"] = "ai_generated"
    scheme_data["title"] = f"{subject} — Year {year_group} ({key_stage})"
    scheme_data["generation"] = {
        "model": model,
        "duration_s": duration,
        "usage": {
            "input_tokens": response.usage.input_tokens,
            "output_tokens": response.usage.output_tokens,
        },
    }

    return save_scheme(scheme_data)


# ───────────────────────────────────────────────────────────────────
# Apply suggestion
# ───────────────────────────────────────────────────────────────────

def apply_suggestion(scheme_id, suggestion, reference_context="",
                     model="claude-sonnet-4-5-20250929"):
    """Apply a single review suggestion to a scheme, returning updated scheme."""
    scheme = get_scheme(scheme_id)
    if not scheme:
        raise ValueError(f"Scheme not found: {scheme_id}")

    start = time.time()

    user_prompt = f"""Apply this suggestion to the scheme of work:

SUGGESTION:
{json.dumps(suggestion, indent=2)}

CURRENT SCHEME:
{json.dumps(scheme, indent=2)}

Return the COMPLETE modified scheme as JSON (same structure, with the suggestion applied).
Do NOT remove or change anything else — only apply the requested change."""

    response = _create_message(
        model=model,
        max_tokens=16000,
        system=[{
            "type": "text",
            "text": SOW_APPLY_SUGGESTION_PROMPT,
            "cache_control": {"type": "ephemeral"},
        }],
        messages=[{"role": "user", "content": user_prompt}],
    )

    duration = round(time.time() - start, 1)
    raw_text = response.content[0].text.strip()

    try:
        updated = _repair_json(raw_text, expect_object=True)
    except json.JSONDecodeError as e:
        raise ValueError(f"AI returned invalid JSON for updated scheme: {e}")

    # Preserve metadata
    updated["id"] = scheme["id"]
    updated["created_at"] = scheme.get("created_at", "")
    updated["source"] = scheme.get("source", "")
    return save_scheme(updated)


# ───────────────────────────────────────────────────────────────────
# Export to pipeline-compatible Excel
# ───────────────────────────────────────────────────────────────────

def export_to_excel(scheme_id, output_path=None):
    """Export a scheme to multi-sheet Excel matching the exemplar format.

    Sheets produced:
      1. Overview — rationale, lesson structure, design decisions
      2. Year X Lessons — full lesson table with all columns
      3. Disciplinary Progression — skill ladders with 5 levels
      4. Curriculum Map — substantive concept threads
      5. Vocabulary Teaching — pre-teaching, retrieval, re-use guidance
      6. Key — colour legend
    """
    scheme = get_scheme(scheme_id)
    if not scheme:
        raise ValueError(f"Scheme not found: {scheme_id}")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Style definitions ────────────────────────────────────────
    bold = Font(bold=True)
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    wrap = Alignment(wrap_text=True, vertical="top")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text = Font(bold=True, color="FFFFFF", size=10)
    assess_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
    retrieval_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # yellow
    consol_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")  # purple
    writing_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")  # blue
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    subject = scheme.get("subject", "Subject")
    year_group = scheme.get("yearGroup", "")
    key_stage = scheme.get("keyStage", "")

    # ── Sheet 1: Overview ────────────────────────────────────────
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.column_dimensions["A"].width = 120

    overview = scheme.get("overview", {})
    ws_overview.append([overview.get("title", f"{subject} {key_stage} Scheme of Work")])
    ws_overview["A1"].font = title_font
    ws_overview.append([overview.get("subtitle", f"Year {year_group} | {scheme.get('totalLessons', '')} lessons")])
    ws_overview["A2"].font = header_font
    ws_overview.append([])

    # Lesson structure
    ws_overview.append(["LESSON STRUCTURE (every lesson):"])
    ws_overview[f"A{ws_overview.max_row}"].font = bold
    for item in overview.get("lessonStructure", []):
        ws_overview.append([item])

    ws_overview.append([])
    ws_overview.append(["DESIGN RATIONALE:"])
    ws_overview[f"A{ws_overview.max_row}"].font = bold
    for item in overview.get("designRationale", []):
        ws_overview.append([item])

    ws_overview.append([])
    ws_overview.append(["CONTENT DECISIONS:"])
    ws_overview[f"A{ws_overview.max_row}"].font = bold
    for item in overview.get("contentDecisions", []):
        ws_overview.append([item])

    ws_overview.append([])
    assessment_model = overview.get("assessmentModel", "")
    if assessment_model:
        ws_overview.append([f"ASSESSMENT: {assessment_model}"])
        ws_overview[f"A{ws_overview.max_row}"].font = bold
    consol_model = overview.get("consolidationModel", "")
    if consol_model:
        ws_overview.append([f"CONSOLIDATION: {consol_model}"])
        ws_overview[f"A{ws_overview.max_row}"].font = bold

    # Apply wrap to all overview cells
    for row in ws_overview.iter_rows():
        for cell in row:
            cell.alignment = wrap

    # ── Sheet 2: Year X Lessons ──────────────────────────────────
    ws_lessons = wb.create_sheet(f"Year_{year_group}_Lessons")

    # Title rows
    ws_lessons.append([f"Year {year_group} Lessons"])
    ws_lessons["A1"].font = title_font
    ws_lessons.append([overview.get("subtitle", f"Year {year_group} | {scheme.get('totalLessons', '')} lessons")])
    ws_lessons["A2"].font = header_font

    # Column headers
    lesson_headers = [
        "Wk", "L#", "Enquiry / Unit", "Lesson Title",
        "Specification Content & Learning Objectives",
        "Key Vocabulary (Pre-taught)",
        "Skill / Concept", "Progression Level",
        "Substantive Concepts", "Retrieval Starter Focus", "Type"
    ]
    ws_lessons.append(lesson_headers)
    for col_idx, _ in enumerate(lesson_headers, 1):
        cell = ws_lessons.cell(row=3, column=col_idx)
        cell.font = header_text
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin_border

    # Column widths
    col_widths = [5, 5, 30, 30, 60, 30, 18, 30, 25, 25, 5]
    for i, w in enumerate(col_widths, 1):
        ws_lessons.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Write lesson rows — support both new "units" and legacy "terms" structure
    units_data = scheme.get("units", [])
    if not units_data:
        # Fallback to legacy "terms" format
        for term in scheme.get("terms", []):
            for unit in term.get("units", []):
                for lesson in unit.get("lessons", []):
                    row_data = [
                        "",
                        lesson.get("number", ""),
                        unit.get("title", ""),
                        lesson.get("title", ""),
                        "; ".join(lesson.get("objectives", [])),
                        ", ".join(lesson.get("keyVocabulary", [])),
                        "", "", "", "", ""
                    ]
                    ws_lessons.append(row_data)
    else:
        for unit_block in units_data:
            for lesson in unit_block.get("lessons", []):
                row_data = [
                    lesson.get("week", ""),
                    lesson.get("number", ""),
                    lesson.get("unit", unit_block.get("enquiryTitle", "")),
                    lesson.get("title", ""),
                    lesson.get("content", ""),
                    lesson.get("vocabulary", ""),
                    lesson.get("skill", ""),
                    lesson.get("progressionLevel", ""),
                    lesson.get("concepts", ""),
                    lesson.get("retrievalFocus", ""),
                    lesson.get("lessonType", ""),
                ]
                row_idx = ws_lessons.max_row + 1
                ws_lessons.append(row_data)

                # Apply colour based on lesson type
                lt = lesson.get("lessonType", "")
                fill = None
                if lt == "A":
                    fill = assess_fill
                elif lt == "C":
                    fill = consol_fill
                elif lt == "W":
                    fill = writing_fill

                if fill:
                    for col_idx in range(1, len(lesson_headers) + 1):
                        ws_lessons.cell(row=row_idx, column=col_idx).fill = fill

    # Apply formatting to all lesson data rows
    for row in ws_lessons.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = wrap
            cell.border = thin_border

    # ── Sheet 3: Disciplinary Progression ────────────────────────
    ws_prog = wb.create_sheet("Disciplinary Progression")
    ws_prog.append(["Subject-Specific Skill Progression Ladders"])
    ws_prog["A1"].font = title_font

    prog_ladders = scheme.get("progressionLadders", [])
    for ladder in prog_ladders:
        ws_prog.append([])
        skill_name = ladder.get("skill", "Skill")
        ws_prog.append([skill_name, skill_name, skill_name, skill_name, skill_name])
        ws_prog.cell(row=ws_prog.max_row, column=1).font = Font(bold=True, size=12)

        expectations = ladder.get("yearExpectations", "")
        if expectations:
            ws_prog.append([expectations] * 5)

        ws_prog.append(["Level", "Descriptor", "Worked Example from Scheme",
                        "Common Misconception", "Pupil-Friendly Version"])
        for col_idx in range(1, 6):
            cell = ws_prog.cell(row=ws_prog.max_row, column=col_idx)
            cell.font = bold
            cell.fill = header_fill
            cell.font = header_text
            cell.border = thin_border

        for level in ladder.get("levels", []):
            ws_prog.append([
                f"{level.get('level', '')}: {level.get('name', '')}",
                level.get("descriptor", ""),
                level.get("workedExample", ""),
                level.get("misconception", ""),
                level.get("pupilFriendly", ""),
            ])
            for col_idx in range(1, 6):
                cell = ws_prog.cell(row=ws_prog.max_row, column=col_idx)
                cell.alignment = wrap
                cell.border = thin_border

    # Column widths for progression
    for i, w in enumerate([25, 40, 50, 50, 40], 1):
        ws_prog.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Sheet 4: Curriculum Map ──────────────────────────────────
    ws_map = wb.create_sheet("Curriculum Map")
    ws_map.append([f"Substantive Concept Threads — {subject} Year {year_group}"])
    ws_map["A1"].font = title_font
    ws_map.append([])

    ws_map.append(["Concept Thread", "Coverage Across the Year"])
    for col_idx in range(1, 3):
        cell = ws_map.cell(row=3, column=col_idx)
        cell.font = header_text
        cell.fill = header_fill
        cell.border = thin_border

    for thread in scheme.get("curriculumMap", []):
        ws_map.append([
            thread.get("thread", ""),
            thread.get("coverage", ""),
        ])
        for col_idx in range(1, 3):
            cell = ws_map.cell(row=ws_map.max_row, column=col_idx)
            cell.alignment = wrap
            cell.border = thin_border

    ws_map.column_dimensions["A"].width = 30
    ws_map.column_dimensions["B"].width = 80

    # ── Sheet 5: Vocabulary Teaching ─────────────────────────────
    ws_vocab = wb.create_sheet("Vocabulary Teaching")
    ws_vocab.append(["Vocabulary Teaching Guidance & Routines"])
    ws_vocab["A1"].font = title_font
    ws_vocab.append([])

    vocab_data = scheme.get("vocabularyTeaching", {})

    # Pre-teaching routine
    ws_vocab.append(["1. PRE-TEACHING ROUTINE"])
    ws_vocab.cell(row=ws_vocab.max_row, column=1).font = bold
    ws_vocab.append(["Step", "What Teacher Does", "What Pupils Do", "Time"])
    for col_idx in range(1, 5):
        cell = ws_vocab.cell(row=ws_vocab.max_row, column=col_idx)
        cell.font = header_text
        cell.fill = header_fill
        cell.border = thin_border

    for step in vocab_data.get("preTeachingRoutine", []):
        ws_vocab.append([
            step.get("step", ""),
            step.get("teacher", ""),
            step.get("pupil", ""),
            step.get("time", ""),
        ])

    ws_vocab.append([])
    ws_vocab.append(["2. CUMULATIVE VOCABULARY IN RETRIEVAL STARTERS"])
    ws_vocab.cell(row=ws_vocab.max_row, column=1).font = bold
    ws_vocab.append(["Question Type", "Example", "When to Use", "Notes"])
    for col_idx in range(1, 5):
        cell = ws_vocab.cell(row=ws_vocab.max_row, column=col_idx)
        cell.font = header_text
        cell.fill = header_fill
        cell.border = thin_border

    for qt in vocab_data.get("retrievalQuestionTypes", []):
        ws_vocab.append([
            qt.get("type", ""),
            qt.get("example", ""),
            qt.get("frequency", ""),
            qt.get("notes", ""),
        ])

    ws_vocab.append([])
    ws_vocab.append(["3. DELIBERATE RE-USE ACROSS UNITS"])
    ws_vocab.cell(row=ws_vocab.max_row, column=1).font = bold
    ws_vocab.append(["Term", "Introduced", "Re-used in", "Why This Matters"])
    for col_idx in range(1, 5):
        cell = ws_vocab.cell(row=ws_vocab.max_row, column=col_idx)
        cell.font = header_text
        cell.fill = header_fill
        cell.border = thin_border

    for reuse in vocab_data.get("deliberateReuse", []):
        ws_vocab.append([
            reuse.get("term", ""),
            reuse.get("introduced", ""),
            reuse.get("reusedIn", ""),
            reuse.get("why", ""),
        ])

    for w, i in [(30, 1), (50, 2), (50, 3), (40, 4)]:
        ws_vocab.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Sheet 6: Key ─────────────────────────────────────────────
    ws_key = wb.create_sheet("Key")
    ws_key.append(["Colour Key"])
    ws_key["A1"].font = title_font
    ws_key.append([])

    key_items = [
        (assess_fill, "Assessment / DIRT lesson (green)", "A"),
        (retrieval_fill, "Retrieval / Review lesson (yellow)", ""),
        (consol_fill, "Consolidation week lesson (purple)", "C"),
        (writing_fill, "Writing / Skill instruction lesson (blue)", "W"),
    ]
    for fill, desc, code in key_items:
        row_idx = ws_key.max_row + 1
        ws_key.append(["", desc])
        ws_key.cell(row=row_idx, column=1).fill = fill
        ws_key.column_dimensions["A"].width = 5
        ws_key.column_dimensions["B"].width = 50

    # ── Save ─────────────────────────────────────────────────────
    if not output_path:
        output_dir = Path(__file__).parent / "output" / "schemes"
        output_dir.mkdir(parents=True, exist_ok=True)
        safe_title = re.sub(r"[^\w\s-]", "", scheme.get("title", "scheme")).replace(" ", "_")
        output_path = str(output_dir / f"{safe_title}.xlsx")

    wb.save(output_path)
    return output_path
