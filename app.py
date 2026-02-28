"""
Flask app for the booklet production pipeline.

Browse parsed lessons, filter by subject/year/topic, view lesson details,
generate booklets via Claude API, validate, and upload to Google Drive.

Supports multiple courses via scheme-of-work upload.
"""

import json
import logging
import os
import re
import time
import uuid
from datetime import datetime
from pathlib import Path
from threading import Lock

from flask import Flask, Response, render_template, request, jsonify, send_file
from prompt_generator import generate_master_prompt
import courses
import tracker

app = Flask(__name__)

# Logging
logging.basicConfig(
    filename="app.log",
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger(__name__)

# Directory for uploaded spreadsheets
UPLOADS_DIR = Path(__file__).parent / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

# Parsed data cache: { course_id: parsed_data_dict }
_course_data = {}

# Current active course
_active_course_id = courses.get_default_course_id()


def get_data(course_id=None):
    """Get parsed data for a course, loading/parsing on first access."""
    global _active_course_id
    cid = course_id or _active_course_id

    if cid not in _course_data:
        config = courses.get_course(cid)
        if not config:
            raise ValueError(f"Course not found: {cid}")

        # Use the generic parser for all courses
        from generic_parser import parse_course
        _course_data[cid] = parse_course(config)

    return _course_data[cid]


def get_active_course_config():
    """Get the full config for the currently active course."""
    return courses.get_course(_active_course_id)


# --- Batch generation state ---
_batch_state = {}
_batch_lock = Lock()


# ========================================================================
# Pages
# ========================================================================

@app.route("/")
def index():
    data = get_data()
    stats = data["stats"]
    subjects = sorted(set(
        l["subject"] for l in data["booklet_lessons"] if l["subject"]
    ))
    topics = sorted(set(
        l["topic"] for l in data["booklet_lessons"] if l["topic"]
    ))
    years = sorted(set(
        l["year"] for l in data["booklet_lessons"]
    ))
    course_list = courses.list_courses()
    active_config = get_active_course_config()
    return render_template(
        "index.html",
        stats=stats,
        subjects=subjects,
        topics=topics,
        years=years,
        courses=course_list,
        active_course=active_config,
    )


# ========================================================================
# Course management API
# ========================================================================

@app.route("/api/courses")
def api_courses():
    return jsonify({
        "courses": courses.list_courses(),
        "active": _active_course_id,
    })


@app.route("/api/courses/switch", methods=["POST"])
def api_switch_course():
    global _active_course_id
    body = request.get_json() or {}
    course_id = body.get("course_id")
    if not course_id:
        return jsonify({"error": "course_id required"}), 400

    config = courses.get_course(course_id)
    if not config:
        return jsonify({"error": "Course not found"}), 404

    _active_course_id = course_id
    # Force re-parse on next access
    _course_data.pop(course_id, None)

    return jsonify({"active": course_id, "name": config["name"]})


@app.route("/api/courses/<course_id>")
def api_get_course(course_id):
    config = courses.get_course(course_id)
    if not config:
        return jsonify({"error": "Course not found"}), 404
    return jsonify(config)


@app.route("/api/courses/<course_id>", methods=["DELETE"])
def api_delete_course(course_id):
    try:
        courses.delete_course(course_id)
        _course_data.pop(course_id, None)
        return jsonify({"deleted": True})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/courses/upload", methods=["POST"])
def api_upload_spreadsheet():
    """Upload a scheme of work spreadsheet and preview its structure."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx or .xls files are supported"}), 400

    # Save the uploaded file
    safe_name = file.filename.replace(" ", "_")
    save_path = UPLOADS_DIR / safe_name
    file.save(str(save_path))

    # Preview the spreadsheet
    from generic_parser import preview_spreadsheet
    try:
        preview = preview_spreadsheet(str(save_path))
    except Exception as e:
        return jsonify({"error": f"Could not read spreadsheet: {e}"}), 400

    return jsonify({
        "filename": safe_name,
        "path": str(save_path),
        "preview": preview,
    })


@app.route("/api/courses/preview", methods=["POST"])
def api_preview_spreadsheet():
    """Preview an already-uploaded spreadsheet (or re-preview with different settings)."""
    body = request.get_json() or {}
    xlsx_path = body.get("xlsx_path")
    if not xlsx_path:
        return jsonify({"error": "xlsx_path required"}), 400

    from generic_parser import preview_spreadsheet
    try:
        preview = preview_spreadsheet(xlsx_path)
    except Exception as e:
        return jsonify({"error": f"Could not read spreadsheet: {e}"}), 400

    return jsonify({"preview": preview})


@app.route("/api/courses/save", methods=["POST"])
def api_save_course():
    """Save a new course configuration after the user has mapped columns."""
    body = request.get_json() or {}

    required = ["name", "xlsx_path", "sheets", "col_map"]
    for field in required:
        if field not in body:
            return jsonify({"error": f"Missing required field: {field}"}), 400

    config = {
        "name": body["name"],
        "exam_board": body.get("exam_board", ""),
        "qualification": body.get("qualification", ""),
        "key_stage": body.get("key_stage", ""),
        "subjects": body.get("subjects", []),
        "xlsx_path": body["xlsx_path"],
        "sheets": body["sheets"],
        "header_row": body.get("header_row", 1),
        "col_map": body["col_map"],
        "topic_folders": body.get("topic_folders", {}),
        "subject_from_topic_prefix": body.get("subject_from_topic_prefix", {}),
        "topic_code_pattern": body.get("topic_code_pattern", ""),
        "fixed_subject": body.get("fixed_subject"),
        "system_prompt_context": body.get("system_prompt_context", body["name"]),
        "prior_knowledge_base": body.get(
            "prior_knowledge_base",
            f"Students have completed the relevant prior learning for this course."
        ),
    }

    # If editing an existing course, keep its ID
    if body.get("id"):
        config["id"] = body["id"]

    saved = courses.save_course(config)

    # Clear cache so it re-parses
    _course_data.pop(saved["id"], None)

    return jsonify({"saved": True, "course": saved})


# ========================================================================
# Update Scheme of Work API
# ========================================================================

@app.route("/api/courses/<course_id>/update-scheme", methods=["POST"])
def api_update_scheme_preview(course_id):
    """Upload a new spreadsheet and preview what will change (non-destructive)."""
    config = courses.get_course(course_id)
    if not config:
        return jsonify({"error": "Course not found"}), 404

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx or .xls files supported"}), 400

    safe_name = file.filename.replace(" ", "_")
    save_path = UPLOADS_DIR / safe_name
    file.save(str(save_path))

    from generic_parser import parse_course, compare_schemes
    import openpyxl

    new_config = dict(config)
    new_config["xlsx_path"] = str(save_path)

    # ── Auto-remap sheet names if they changed ──
    try:
        wb = openpyxl.load_workbook(str(save_path), read_only=True)
        actual_sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        return jsonify({"error": f"Could not open spreadsheet: {e}"}), 400

    sheet_remap = {}
    new_sheets = list(new_config.get("sheets", []))
    for i, sheet_info in enumerate(new_sheets):
        old_name = sheet_info["name"]
        if old_name in actual_sheets:
            continue  # exact match, no remap needed
        year = sheet_info["year"]
        # Try fuzzy match: find a sheet containing the year number
        # and a keyword like "lesson" (case-insensitive)
        candidates = []
        for sn in actual_sheets:
            sn_lower = sn.lower().replace("_", " ")
            if str(year) in sn and "lesson" in sn_lower:
                candidates.append(sn)
        if len(candidates) == 1:
            sheet_remap[old_name] = candidates[0]
            new_sheets[i] = {**sheet_info, "name": candidates[0]}
        elif not candidates:
            # Broader match: just contains the year number
            for sn in actual_sheets:
                if str(year) in sn and sn not in [s["name"] for s in new_sheets[:i]]:
                    candidates.append(sn)
            if len(candidates) == 1:
                sheet_remap[old_name] = candidates[0]
                new_sheets[i] = {**sheet_info, "name": candidates[0]}

    new_config["sheets"] = new_sheets

    # ── Auto-detect column changes ──
    # Check if header row columns shifted by comparing expected headers
    col_warnings = []
    for sheet_info in new_sheets:
        sn = sheet_info["name"]
        if sn not in actual_sheets:
            continue
        try:
            wb2 = openpyxl.load_workbook(str(save_path), read_only=True)
            ws = wb2[sn]
            header_row_idx = new_config.get("header_row", 1)
            for ri, row in enumerate(ws.iter_rows(max_row=header_row_idx + 1, values_only=True), 1):
                if ri == header_row_idx + 1:
                    headers = [str(c).strip().lower() if c else "" for c in row]
            wb2.close()

            col_map = new_config.get("col_map", {})
            # Check that key columns still look right
            for field, idx in col_map.items():
                if idx < len(headers):
                    h = headers[idx]
                    # Basic sanity: if 'topic' column now says 'disciplinary' or 'progression', flag it
                    if field == "topic" and ("disciplin" in h or "progression" in h):
                        # Try to find the actual topic/substantive column
                        for ni, nh in enumerate(headers):
                            if "substantive" in nh or "concept" in nh and "second" not in nh:
                                col_warnings.append(f"Column '{field}' remapped from {idx} to {ni} (was '{headers[idx]}', now '{nh}')")
                                new_config.setdefault("col_map", {})[field] = ni
                                break
        except Exception:
            pass

    try:
        new_data = parse_course(new_config)
    except Exception as e:
        return jsonify({"error": f"Could not parse spreadsheet: {e}"}), 400

    try:
        old_data = get_data(course_id)
    except Exception:
        old_data = {"booklet_lessons": [], "all_lessons": []}

    diff = compare_schemes(
        old_data["booklet_lessons"],
        new_data["booklet_lessons"]
    )

    return jsonify({
        "new_xlsx_path": str(save_path),
        "diff": {
            "summary": diff["summary"],
            "modified": [
                {
                    "year": m["key"][0],
                    "lesson_number": m["key"][1],
                    "old_title": m["old"].get("title", ""),
                    "new_title": m["new"].get("title", ""),
                }
                for m in diff["modified"]
            ],
            "added": [
                {
                    "year": a["key"][0],
                    "lesson_number": a["key"][1],
                    "title": a["new"].get("title", ""),
                }
                for a in diff["added"]
            ],
            "removed": [
                {
                    "year": r["key"][0],
                    "lesson_number": r["key"][1],
                    "title": r["old"].get("title", ""),
                }
                for r in diff["removed"]
            ],
        },
        "new_stats": new_data["stats"],
        "sheet_remap": sheet_remap,
        "col_remap": {k: v for k, v in new_config.get("col_map", {}).items()},
        "col_warnings": col_warnings,
        "new_sheets": new_sheets,
    })


@app.route("/api/courses/<course_id>/apply-update", methods=["POST"])
def api_apply_scheme_update(course_id):
    """Apply a previously previewed scheme update (destructive, user-confirmed)."""
    config = courses.get_course(course_id)
    if not config:
        return jsonify({"error": "Course not found"}), 404

    body = request.get_json() or {}
    new_xlsx_path = body.get("new_xlsx_path")
    delete_removed = body.get("delete_removed", True)
    reset_modified = body.get("reset_modified", True)
    delete_from_drive = body.get("delete_from_drive", False)
    new_sheets = body.get("new_sheets")  # remapped sheet names from preview
    col_remap = body.get("col_remap")    # remapped column indices from preview

    if not new_xlsx_path:
        return jsonify({"error": "new_xlsx_path required"}), 400

    from generic_parser import parse_course, compare_schemes
    from generator import delete_lesson_files_from_disk

    try:
        old_data = get_data(course_id)
    except Exception:
        old_data = {"booklet_lessons": [], "all_lessons": []}

    new_config = dict(config)
    new_config["xlsx_path"] = new_xlsx_path
    if new_sheets:
        new_config["sheets"] = new_sheets
    if col_remap:
        # col_remap values come as strings from JSON; convert to int
        new_config["col_map"] = {k: int(v) for k, v in col_remap.items()}
    new_data = parse_course(new_config)

    diff = compare_schemes(
        old_data["booklet_lessons"],
        new_data["booklet_lessons"]
    )

    results = {
        "disk_deleted": [],
        "drive_deleted": [],
        "progress_cleared": [],
        "errors": [],
    }

    # Handle removed lessons
    if delete_removed and diff["removed"]:
        for item in diff["removed"]:
            lesson = item["old"]
            try:
                disk_result = delete_lesson_files_from_disk(lesson)
                results["disk_deleted"].extend(disk_result["deleted"])
            except Exception as e:
                results["errors"].append(
                    f"Disk delete failed for Y{lesson['year']}_L{lesson['lesson_number']}: {e}"
                )

            if delete_from_drive:
                try:
                    from gdrive import delete_lesson_files_from_drive
                    drive_folder = config.get("gdrive_folder_id") or None
                    drive_result = delete_lesson_files_from_drive(
                        lesson, root_folder_id=drive_folder
                    )
                    results["drive_deleted"].extend(drive_result["deleted"])
                    results["errors"].extend(drive_result["errors"])
                except Exception as e:
                    results["errors"].append(
                        f"Drive delete failed for Y{lesson['year']}_L{lesson['lesson_number']}: {e}"
                    )

        removed_keys = [
            (item["old"]["year"], item["old"]["lesson_number"])
            for item in diff["removed"]
        ]
        tracker.clear_lessons(removed_keys, course_id)
        results["progress_cleared"].extend(
            [f"Y{y}_L{n:03d}" for y, n in removed_keys]
        )

    # Handle modified lessons — reset progress so they get regenerated
    if reset_modified and diff["modified"]:
        modified_keys = [(m["key"][0], m["key"][1]) for m in diff["modified"]]
        tracker.clear_lessons(modified_keys, course_id)
        results["progress_cleared"].extend(
            [f"Y{y}_L{n:03d}" for y, n in modified_keys]
        )

    # Update course config with new spreadsheet path, sheets, and col_map
    config["xlsx_path"] = new_xlsx_path
    if new_sheets:
        config["sheets"] = new_sheets
    if col_remap:
        config["col_map"] = {k: int(v) for k, v in col_remap.items()}
    courses.save_course(config)

    # Clear cache to force re-parse
    _course_data.pop(course_id, None)

    try:
        get_data(course_id)
    except Exception as e:
        results["errors"].append(f"Re-parse failed: {e}")

    return jsonify({
        "success": True,
        "results": results,
        "diff_summary": diff["summary"],
    })


# ========================================================================
# Export Scheme of Work API
# ========================================================================

@app.route("/api/courses/<course_id>/export-scheme", methods=["POST"])
def api_export_scheme(course_id):
    """Export the scheme of work in various formats (download or Google Drive)."""
    config = courses.get_course(course_id)
    if not config:
        return jsonify({"error": "Course not found"}), 404

    body = request.get_json() or {}
    fmt = body.get("format", "xlsx")
    course_name = config.get("name", "Scheme")

    if fmt == "xlsx":
        xlsx_path = Path(config["xlsx_path"])
        if not xlsx_path.exists():
            return jsonify({"error": "Source spreadsheet not found"}), 404
        return send_file(
            str(xlsx_path),
            as_attachment=True,
            download_name=f"{course_name} - Scheme of Work.xlsx",
        )

    elif fmt == "docx":
        from scheme_export import export_scheme_docx
        data = get_data(course_id)
        docx_path = export_scheme_docx(config, data["all_lessons"])
        return send_file(
            docx_path,
            as_attachment=True,
            download_name=f"{course_name} - Scheme of Work.docx",
        )

    elif fmt == "pdf":
        from scheme_export import export_scheme_docx
        from generator import convert_to_pdf
        data = get_data(course_id)
        docx_path = export_scheme_docx(config, data["all_lessons"])
        pdf_path = convert_to_pdf(docx_path)
        if not pdf_path:
            return jsonify({"error": "PDF conversion failed — is LibreOffice installed?"}), 500
        return send_file(
            str(pdf_path),
            as_attachment=True,
            download_name=f"{course_name} - Scheme of Work.pdf",
        )

    elif fmt == "google_sheets":
        from gdrive import upload_as_google_native
        xlsx_path = config["xlsx_path"]
        result = upload_as_google_native(
            xlsx_path,
            f"{course_name} - Scheme of Work",
            target_mime="application/vnd.google-apps.spreadsheet",
            folder_id=config.get("gdrive_folder_id") or None,
        )
        return jsonify(result)

    elif fmt == "google_docs":
        from scheme_export import export_scheme_docx
        from gdrive import upload_as_google_native
        data = get_data(course_id)
        docx_path = export_scheme_docx(config, data["all_lessons"])
        result = upload_as_google_native(
            docx_path,
            f"{course_name} - Scheme of Work",
            target_mime="application/vnd.google-apps.document",
            folder_id=config.get("gdrive_folder_id") or None,
        )
        return jsonify(result)

    else:
        return jsonify({"error": f"Unknown format: {fmt}"}), 400


@app.route("/api/courses/<course_id>/save-scheme", methods=["POST"])
def api_save_scheme(course_id):
    """Save the scheme of work to the Mac output folder."""
    from generator import OUTPUT_DIR
    config = courses.get_course(course_id)
    if not config:
        return jsonify({"error": "Course not found"}), 404

    body = request.get_json() or {}
    fmt = body.get("format", "xlsx")
    course_name = config.get("name", "Scheme")

    # Create output folder: output/{course_name}/Scheme of Work/
    from generic_parser import _sanitize_folder_name
    safe_name = _sanitize_folder_name(course_name)
    save_dir = OUTPUT_DIR / safe_name / "Scheme of Work"
    save_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{course_name} - Scheme of Work"

    if fmt == "xlsx":
        import shutil
        src = Path(config["xlsx_path"])
        if not src.exists():
            return jsonify({"error": "Source spreadsheet not found"}), 404
        dest = save_dir / f"{filename}.xlsx"
        shutil.copy2(str(src), str(dest))
        return jsonify({"success": True, "path": str(dest)})

    elif fmt == "docx":
        from scheme_export import export_scheme_docx
        data = get_data(course_id)
        tmp_path = export_scheme_docx(config, data["all_lessons"])
        dest = save_dir / f"{filename}.docx"
        import shutil
        shutil.move(tmp_path, str(dest))
        return jsonify({"success": True, "path": str(dest)})

    elif fmt == "pdf":
        from scheme_export import export_scheme_docx
        from generator import convert_to_pdf
        data = get_data(course_id)
        tmp_path = export_scheme_docx(config, data["all_lessons"])
        pdf_path = convert_to_pdf(tmp_path)
        if not pdf_path:
            return jsonify({"error": "PDF conversion failed — is LibreOffice installed?"}), 500
        dest = save_dir / f"{filename}.pdf"
        import shutil
        shutil.move(str(pdf_path), str(dest))
        # Clean up the temp docx
        Path(tmp_path).unlink(missing_ok=True)
        return jsonify({"success": True, "path": str(dest)})

    else:
        return jsonify({"error": f"Unknown format: {fmt}"}), 400


# ========================================================================
# Lessons API
# ========================================================================

@app.route("/api/lessons")
def api_lessons():
    data = get_data()
    lessons = data["booklet_lessons"]

    subject = request.args.get("subject")
    year = request.args.get("year")
    topic = request.args.get("topic")
    search = request.args.get("search", "").strip().lower()

    if subject:
        lessons = [l for l in lessons if l["subject"] == subject]
    if year:
        lessons = [l for l in lessons if str(l["year"]) == year]
    if topic:
        lessons = [l for l in lessons if l["topic"] == topic]
    if search:
        lessons = [
            l for l in lessons
            if search in (l["title"] or "").lower()
            or search in (l["spec_content"] or "").lower()
            or search in (l["key_vocabulary"] or "").lower()
        ]

    status_filter = request.args.get("status")
    if status_filter:
        statuses = tracker.get_all_statuses(_active_course_id)
        lessons = [
            l for l in lessons
            if statuses.get(tracker._key(l['year'], l['lesson_number'], _active_course_id), "pending") == status_filter
        ]

    statuses = tracker.get_all_statuses(_active_course_id)
    for l in lessons:
        l["status"] = statuses.get(tracker._key(l['year'], l['lesson_number'], _active_course_id), "pending")

    return jsonify(lessons)


@app.route("/api/lesson/<int:year>/<int:lesson_num>")
def api_lesson_detail(year, lesson_num):
    data = get_data()
    for l in data["all_lessons"]:
        if l["year"] == year and l["lesson_number"] == lesson_num:
            return jsonify(l)
    return jsonify({"error": "Lesson not found"}), 404


@app.route("/api/prompt/<int:year>/<int:lesson_num>")
def api_prompt(year, lesson_num):
    data = get_data()
    for l in data["booklet_lessons"]:
        if l["year"] == year and l["lesson_number"] == lesson_num:
            config = get_active_course_config()
            prompt = generate_master_prompt(l, course_config=config)
            return jsonify({"prompt": prompt, "lesson": l})
    return jsonify({"error": "Lesson not found"}), 404


@app.route("/api/filtered-out")
def api_filtered_out():
    data = get_data()
    non_booklet = [l for l in data["all_lessons"] if not l["is_booklet_lesson"]]
    return jsonify(non_booklet)


@app.route("/api/reload", methods=["POST"])
def api_reload():
    global _course_data
    _course_data.pop(_active_course_id, None)
    get_data()
    return jsonify({"status": "ok", "stats": get_data()["stats"]})


# ========================================================================
# Progress tracking
# ========================================================================

@app.route("/api/status/<int:year>/<int:lesson_num>")
def api_get_status(year, lesson_num):
    return jsonify({"status": tracker.get_status(year, lesson_num, _active_course_id)})


@app.route("/api/status/<int:year>/<int:lesson_num>", methods=["POST"])
def api_set_status(year, lesson_num):
    body = request.get_json()
    status = body.get("status", "pending")
    try:
        tracker.set_status(year, lesson_num, status, _active_course_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"status": status})


@app.route("/api/statuses")
def api_all_statuses():
    return jsonify(tracker.get_all_statuses(_active_course_id))


@app.route("/api/progress-summary")
def api_progress_summary():
    data = get_data()
    return jsonify(tracker.get_summary(data["booklet_lessons"], _active_course_id))


# ========================================================================
# Batch export (prompts to .txt)
# ========================================================================

EXPORT_DIR = Path(__file__).parent / "prompts"


@app.route("/api/export", methods=["POST"])
def api_export_prompts():
    data = get_data()
    lessons = data["booklet_lessons"]
    config = get_active_course_config()

    body = request.get_json() or {}
    subject = body.get("subject")
    year = body.get("year")
    status_filter = body.get("status")

    if subject:
        lessons = [l for l in lessons if l["subject"] == subject]
    if year:
        lessons = [l for l in lessons if str(l["year"]) == str(year)]
    if status_filter:
        statuses = tracker.get_all_statuses(_active_course_id)
        lessons = [
            l for l in lessons
            if statuses.get(tracker._key(l['year'], l['lesson_number'], _active_course_id), "pending") == status_filter
        ]

    EXPORT_DIR.mkdir(exist_ok=True)
    exported = []
    for l in lessons:
        prompt = generate_master_prompt(l, course_config=config)
        subdir = EXPORT_DIR / (l["subject"] or "Unknown")
        subdir.mkdir(exist_ok=True)
        fname = f"L{l['lesson_number']:03d} - {l['title']}.txt"
        fname = fname.replace("/", "-").replace(":", " -")
        fpath = subdir / fname
        fpath.write_text(prompt)
        exported.append(str(fpath.relative_to(EXPORT_DIR)))

    return jsonify({
        "exported": len(exported),
        "directory": str(EXPORT_DIR),
        "files": exported,
    })


# ========================================================================
# Check if booklet exists
# ========================================================================

@app.route("/api/check-exists/<int:year>/<int:lesson_num>")
def api_check_exists(year, lesson_num):
    from generator import check_existing_booklet

    data = get_data()
    lesson = _find_lesson(data, year, lesson_num)
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    result = check_existing_booklet(lesson)
    return jsonify(result)


# ========================================================================
# Generate via Claude API
# ========================================================================

@app.route("/api/generate/<int:year>/<int:lesson_num>", methods=["POST"])
def api_generate(year, lesson_num):
    from generator import generate_and_save

    data = get_data()
    lesson = _find_lesson(data, year, lesson_num)
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    body = request.get_json() or {}
    model = body.get("model", "claude-sonnet-4-5-20250929")
    replace = body.get("replace", False)

    config = get_active_course_config()
    prompt = generate_master_prompt(lesson, course_config=config)

    try:
        result = generate_and_save(
            lesson, prompt, model=model, replace=replace,
            course_config=config,
        )

        if replace:
            logger.info(
                f"REPLACED booklet Y{year}_L{lesson_num:03d} "
                f"'{lesson['title']}' at {datetime.now().isoformat()}"
            )

        tracker.set_status(year, lesson_num, "generated", _active_course_id)
        return jsonify({
            "success": True,
            "md_path": result["md_path"],
            "docx_path": result["docx_path"],
            "pdf_path": result.get("pdf_path"),
            "usage": result["usage"],
            "model": result["model"],
            "duration_s": result["duration_s"],
        })
    except FileExistsError as e:
        return jsonify({
            "error": str(e),
            "exists": True,
        }), 409
    except Exception as e:
        logger.error(f"Generate failed for Y{year}_L{lesson_num:03d}: {e}")
        return jsonify({"error": str(e)}), 500


# ========================================================================
# Generate All — SSE stream
# ========================================================================

@app.route("/api/generate-all", methods=["POST"])
def api_generate_all():
    from generator import generate_and_save, check_existing_booklet

    data = get_data()
    body = request.get_json() or {}

    subject = body.get("subject")
    year = body.get("year")
    replace_all = body.get("replace_all", False)
    model = body.get("model", "claude-sonnet-4-5-20250929")

    config = get_active_course_config()

    lesson_ids = body.get("lesson_ids")  # e.g. ["Y7_L001", "Y8_L005"] for custom selection

    lessons = list(data["booklet_lessons"])
    if subject:
        lessons = [l for l in lessons if l["subject"] == subject]
    if year:
        lessons = [l for l in lessons if str(l["year"]) == str(year)]

    # Custom lesson selection overrides subject/year filter
    if lesson_ids:
        id_set = set(lesson_ids)
        lessons = [l for l in lessons if f"Y{l['year']}_L{l['lesson_number']:03d}" in id_set]

    # Only generate pending lessons unless replace_all
    if not replace_all:
        statuses = tracker.get_all_statuses(_active_course_id)
        lessons = [
            l for l in lessons
            if statuses.get(tracker._key(l['year'], l['lesson_number'], _active_course_id), "pending") == "pending"
        ]

    batch_id = str(uuid.uuid4())[:8]
    with _batch_lock:
        _batch_state[batch_id] = {"cancelled": False}

    def generate_stream():
        total = len(lessons)
        generated = 0
        replaced = 0
        skipped = 0
        errors = 0

        yield f"data: {json.dumps({'type': 'start', 'batch_id': batch_id, 'total': total})}\n\n"

        for idx, lesson in enumerate(lessons):
            # Check cancellation
            with _batch_lock:
                if _batch_state.get(batch_id, {}).get("cancelled"):
                    yield f"data: {json.dumps({'type': 'cancelled', 'current': idx, 'total': total})}\n\n"
                    break

            y = lesson["year"]
            ln = lesson["lesson_number"]
            title = lesson["title"]

            yield f"data: {json.dumps({'type': 'progress', 'current': idx + 1, 'total': total, 'title': title, 'year': y, 'lesson_num': ln})}\n\n"

            try:
                existing = check_existing_booklet(lesson)
                is_replace = existing["exists"]

                if is_replace and not replace_all:
                    skipped += 1
                    yield f"data: {json.dumps({'type': 'skipped', 'title': title, 'reason': 'exists'})}\n\n"
                    continue

                prompt = generate_master_prompt(lesson, course_config=config)
                result = generate_and_save(
                    lesson, prompt, model=model, replace=True,
                    course_config=config,
                )
                tracker.set_status(y, ln, "generated", _active_course_id)

                if is_replace:
                    replaced += 1
                    logger.info(
                        f"BATCH REPLACED Y{y}_L{ln:03d} '{title}' "
                        f"at {datetime.now().isoformat()}"
                    )
                else:
                    generated += 1

                yield f"data: {json.dumps({'type': 'done', 'title': title, 'replaced': is_replace, 'duration_s': result['duration_s'], 'usage': result['usage'], 'model': result['model']})}\n\n"

            except Exception as e:
                errors += 1
                logger.error(f"Batch generate failed for Y{y}_L{ln:03d}: {e}")
                yield f"data: {json.dumps({'type': 'error', 'title': title, 'error': str(e)})}\n\n"

        summary = {
            "type": "complete",
            "generated": generated,
            "replaced": replaced,
            "skipped": skipped,
            "errors": errors,
            "total": total,
        }
        yield f"data: {json.dumps(summary)}\n\n"

        # Cleanup batch state
        with _batch_lock:
            _batch_state.pop(batch_id, None)

    return Response(
        generate_stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.route("/api/generate-all/cancel", methods=["POST"])
def api_cancel_batch():
    body = request.get_json() or {}
    batch_id = body.get("batch_id")

    if not batch_id:
        # Cancel all batches
        with _batch_lock:
            for bid in _batch_state:
                _batch_state[bid]["cancelled"] = True
        return jsonify({"cancelled": True})

    with _batch_lock:
        if batch_id in _batch_state:
            _batch_state[batch_id]["cancelled"] = True
            return jsonify({"cancelled": True})

    return jsonify({"error": "Batch not found"}), 404


# ========================================================================
# Rebuild all booklets — re-sanitise markdown + rebuild docx/pdf
# ========================================================================

@app.route("/api/reprocess-all", methods=["POST"])
def api_reprocess_all():
    """
    Re-sanitise every .md file in the output directory and rebuild its
    .docx and .pdf.  Used to apply numbering/formatting fixes to booklets
    that were generated with an older version of the pipeline.
    """
    from generator import sanitize_markdown, markdown_to_docx, convert_to_pdf, OUTPUT_DIR

    def stream():
        md_files = sorted(Path(OUTPUT_DIR).rglob("*.md"))
        total = len(md_files)
        yield f"data: {json.dumps({'type': 'start', 'total': total})}\n\n"

        done = errors = 0
        for i, md_path in enumerate(md_files):
            try:
                content = md_path.read_text(encoding="utf-8")
                clean = sanitize_markdown(content)
                md_path.write_text(clean, encoding="utf-8")

                docx_path = markdown_to_docx(str(md_path))
                convert_to_pdf(docx_path)

                done += 1
                yield f"data: {json.dumps({'type': 'progress', 'current': i + 1, 'total': total, 'file': md_path.name})}\n\n"
            except Exception as e:
                errors += 1
                logger.error(f"Reprocess failed for {md_path.name}: {e}")
                yield f"data: {json.dumps({'type': 'error', 'file': md_path.name, 'error': str(e)})}\n\n"

        yield f"data: {json.dumps({'type': 'complete', 'done': done, 'errors': errors, 'total': total})}\n\n"

    return Response(
        stream(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ========================================================================
# Validate .docx
# ========================================================================

@app.route("/api/validate/<int:year>/<int:lesson_num>", methods=["POST"])
def api_validate(year, lesson_num):
    from validator import validate_docx
    from generator import check_existing_booklet

    data = get_data()
    lesson = _find_lesson(data, year, lesson_num)
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    existing = check_existing_booklet(lesson)
    if not existing["exists"]:
        return jsonify({"error": f"File not found: {existing['docx_path']}. Generate it first."}), 404

    docx_path = Path(existing["docx_path"])

    try:
        result = validate_docx(str(docx_path))
        if result["valid"]:
            tracker.set_status(year, lesson_num, "qa_passed", _active_course_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ========================================================================
# Google Drive upload
# ========================================================================

@app.route("/api/upload/<int:year>/<int:lesson_num>", methods=["POST"])
def api_upload(year, lesson_num):
    from gdrive import upload_booklet
    from generator import check_existing_booklet

    data = get_data()
    lesson = _find_lesson(data, year, lesson_num)
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    existing = check_existing_booklet(lesson)
    if not existing["exists"]:
        return jsonify({"error": f"File not found: {existing['docx_path']}. Generate it first."}), 404

    docx_path = Path(existing["docx_path"])

    try:
        # Use per-course Drive folder if configured, otherwise global
        config = get_active_course_config()
        drive_folder = (config or {}).get("gdrive_folder_id") or None
        result = upload_booklet(str(docx_path), lesson, root_folder_id=drive_folder)
        tracker.set_status(year, lesson_num, "uploaded", _active_course_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/gdrive/check")
def api_gdrive_check():
    from gdrive import check_connection
    return jsonify(check_connection())


# ========================================================================
# Config / setup
# ========================================================================

@app.route("/api/config")
def api_config():
    from dotenv import load_dotenv
    load_dotenv()
    return jsonify({
        "anthropic_key": bool(os.getenv("ANTHROPIC_API_KEY")),
        "gdrive_folder": bool(os.getenv("GDRIVE_ROOT_FOLDER_ID")),
        "gdrive_secrets": Path(
            os.getenv("GOOGLE_CLIENT_SECRETS_FILE", "client_secret.json")
        ).exists() or (Path(__file__).parent / os.getenv("GOOGLE_CLIENT_SECRETS_FILE", "client_secret.json")).exists(),
        "gdrive_token": (Path(__file__).parent / "gdrive_token.json").exists(),
    })


@app.route("/api/env", methods=["GET"])
def api_env_read():
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        return jsonify({"exists": False, "content": ""})
    raw = env_path.read_text()
    return jsonify({"exists": True, "content": raw})


@app.route("/api/env", methods=["POST"])
def api_env_write():
    body = request.get_json() or {}
    content = body.get("content", "")
    env_path = Path(__file__).parent / ".env"
    env_path.write_text(content)
    from dotenv import load_dotenv
    load_dotenv(override=True)
    return jsonify({"saved": True})


# ========================================================================
# Feedback Engine API
# ========================================================================

@app.route("/api/feedback/<course_id>/<int:year>/<int:lesson_num>", methods=["GET"])
def api_get_feedback(course_id, year, lesson_num):
    """Return feedback history for a lesson."""
    from feedback_engine import load_feedback_history
    history = load_feedback_history(course_id, year, lesson_num)
    return jsonify({"history": history})


@app.route("/api/feedback/<course_id>/<int:year>/<int:lesson_num>", methods=["POST"])
def api_apply_feedback(course_id, year, lesson_num):
    """Apply teacher feedback to an existing booklet."""
    from feedback_engine import apply_feedback
    from generator import check_existing_booklet

    data = get_data(course_id)
    lesson = None
    for l in data["all_lessons"]:
        if l["year"] == year and l["lesson_number"] == lesson_num:
            lesson = l
            break
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    existing = check_existing_booklet(lesson)
    if not existing["exists"]:
        return jsonify({
            "error": "No booklet exists for this lesson yet. Generate it first."
        }), 404

    body = request.get_json() or {}
    feedback_text = body.get("feedback_text", "").strip()
    if not feedback_text:
        return jsonify({"error": "feedback_text is required"}), 400

    model = body.get("model", "claude-sonnet-4-5-20250929")

    try:
        result = apply_feedback(
            md_path=existing["docx_path"].replace(".docx", ".md"),
            feedback_text=feedback_text,
            lesson=lesson,
            course_id=course_id,
            model=model,
        )
        return jsonify({"success": True, **result})
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logger.error(f"Feedback apply failed for Y{year}_L{lesson_num:03d}: {e}")
        return jsonify({"error": str(e)}), 500


# ========================================================================
# SEND Student Profiles API
# ========================================================================

@app.route("/api/students", methods=["GET"])
def api_list_students():
    import students
    return jsonify({"students": students.list_students()})


@app.route("/api/students", methods=["POST"])
def api_save_student():
    import students
    body = request.get_json() or {}
    try:
        saved = students.save_student(body)
        return jsonify({"saved": True, "student": saved})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/students/<student_id>", methods=["GET"])
def api_get_student(student_id):
    import students
    s = students.get_student(student_id)
    if not s:
        return jsonify({"error": "Student not found"}), 404
    return jsonify(s)


@app.route("/api/students/<student_id>", methods=["PUT"])
def api_update_student(student_id):
    import students
    body = request.get_json() or {}
    body["id"] = student_id
    try:
        saved = students.save_student(body)
        return jsonify({"saved": True, "student": saved})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/students/<student_id>", methods=["DELETE"])
def api_delete_student(student_id):
    import students
    try:
        students.delete_student(student_id)
        return jsonify({"deleted": True})
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


# ========================================================================
# SEND Personalisation Engine API
# ========================================================================

@app.route("/api/send-booklet/<course_id>/<int:year>/<int:lesson_num>/<student_id>",
           methods=["POST"])
def api_generate_send_booklet(course_id, year, lesson_num, student_id):
    """Generate a SEND-personalised booklet for a specific student."""
    from send_engine import generate_send_booklet
    from generator import check_existing_booklet
    import students as students_mod

    data = get_data(course_id)
    lesson = None
    for l in data["all_lessons"]:
        if l["year"] == year and l["lesson_number"] == lesson_num:
            lesson = l
            break
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    student = students_mod.get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404

    existing = check_existing_booklet(lesson)
    if not existing["exists"]:
        return jsonify({
            "error": "No booklet exists for this lesson yet. Generate the standard booklet first."
        }), 404

    body = request.get_json() or {}
    model = body.get("model", "claude-sonnet-4-5-20250929")
    config = courses.get_course(course_id)

    # Find the markdown path from the docx path
    md_path = existing["docx_path"].replace(".docx", ".md")

    try:
        result = generate_send_booklet(
            original_md_path=md_path,
            student=student,
            lesson=lesson,
            course_config=config,
            model=model,
        )
        return jsonify({"success": True, **result})
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.error(
            f"SEND booklet failed for Y{year}_L{lesson_num:03d} "
            f"student {student_id}: {e}"
        )
        return jsonify({"error": str(e)}), 500


@app.route("/api/send-booklets/<course_id>/<int:year>/<int:lesson_num>",
           methods=["GET"])
def api_list_send_booklets(course_id, year, lesson_num):
    """List all SEND-personalised booklets for a lesson."""
    from generator import check_existing_booklet, OUTPUT_DIR
    import re as _re

    data = get_data(course_id)
    lesson = None
    for l in data["all_lessons"]:
        if l["year"] == year and l["lesson_number"] == lesson_num:
            lesson = l
            break
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    existing = check_existing_booklet(lesson)
    if not existing["exists"]:
        return jsonify({"booklets": []})

    docx_path = Path(existing["docx_path"])
    parent = docx_path.parent
    stem = docx_path.stem

    # Find SEND variants: "{stem} - SEND - {Name}.docx"
    booklets = []
    for f in sorted(parent.glob(f"{stem} - SEND - *.docx")):
        m = _re.search(r" - SEND - (.+)\.docx$", f.name)
        student_name_raw = m.group(1).replace("_", " ") if m else f.stem
        booklets.append({
            "filename": f.name,
            "student_name": student_name_raw,
            "path": str(f),
            "pdf_exists": f.with_suffix(".pdf").exists(),
        })

    return jsonify({"booklets": booklets})


# ========================================================================
# Tiered Access API (Prompt Sheet 12)
# ========================================================================

@app.route("/api/tier")
def api_get_tier():
    from access_tiers import get_tier_info
    return jsonify(get_tier_info())


@app.route("/api/tier", methods=["PUT"])
def api_set_tier():
    from access_tiers import set_tier
    body = request.get_json() or {}
    tier = body.get("tier")
    if not tier:
        return jsonify({"error": "tier is required"}), 400
    try:
        new_tier = set_tier(tier)
        return jsonify({"tier": new_tier})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


# ========================================================================
# Reference Document Library API (Prompt Sheet 12)
# ========================================================================

@app.route("/api/reference-docs")
def api_list_reference_docs():
    from access_tiers import requires_tier
    import reference_library as reflib
    category = request.args.get("category")
    subject = request.args.get("subject")
    key_stage = request.args.get("key_stage")
    docs = reflib.list_documents(category=category, subject=subject, key_stage=key_stage)
    return jsonify({"documents": docs})


@app.route("/api/reference-docs/<doc_id>")
def api_get_reference_doc(doc_id):
    import reference_library as reflib
    doc = reflib.get_document(doc_id)
    if not doc:
        return jsonify({"error": "Document not found"}), 404
    return jsonify(doc)


@app.route("/api/reference-docs/upload", methods=["POST"])
def api_upload_reference_doc():
    from access_tiers import get_tier, TIER_HIERARCHY
    if TIER_HIERARCHY.get(get_tier(), 0) < TIER_HIERARCHY["all_custom"]:
        return jsonify({"error": "upgrade_required", "message": "Upload requires All Custom access."}), 403

    import reference_library as reflib

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    category = request.form.get("category", "expert_input")
    title = request.form.get("title", file.filename)
    subject = request.form.get("subject") or None
    key_stage = request.form.get("key_stage") or None
    exam_board = request.form.get("exam_board") or None
    description = request.form.get("description") or None
    doc_id = request.form.get("doc_id") or None  # for replacement

    # Save uploaded file temporarily
    safe_name = file.filename.replace(" ", "_")
    tmp_path = UPLOADS_DIR / safe_name
    file.save(str(tmp_path))

    try:
        doc = reflib.save_document(
            file_path=str(tmp_path),
            category=category,
            title=title,
            subject=subject,
            key_stage=key_stage,
            exam_board=exam_board,
            description=description,
            doc_id=doc_id,
        )
        return jsonify({"saved": True, "document": doc})
    except (ValueError, FileNotFoundError) as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/reference-docs/<doc_id>", methods=["PUT"])
def api_update_reference_doc(doc_id):
    import reference_library as reflib
    body = request.get_json() or {}
    try:
        doc = reflib.update_document(doc_id, **body)
        return jsonify({"updated": True, "document": doc})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/reference-docs/<doc_id>", methods=["DELETE"])
def api_delete_reference_doc(doc_id):
    import reference_library as reflib
    try:
        reflib.delete_document(doc_id)
        return jsonify({"deleted": True})
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


@app.route("/api/reference-docs/<doc_id>/reparse", methods=["POST"])
def api_reparse_reference_doc(doc_id):
    import reference_library as reflib
    try:
        result = reflib.reparse_document(doc_id)
        return jsonify(result)
    except (ValueError, FileNotFoundError) as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/reference-docs/context")
def api_reference_context():
    import reference_library as reflib
    subject = request.args.get("subject")
    key_stage = request.args.get("key_stage")
    exam_board = request.args.get("exam_board")
    context = reflib.get_reference_context(subject=subject, key_stage=key_stage, exam_board=exam_board)
    return jsonify({"context": context, "length": len(context)})


@app.route("/api/reference-docs/update-thinkers", methods=["POST"])
def api_update_thinkers():
    """Regenerate key thinker expert-input files via AI and re-import into library."""
    import threading
    from generate_expert_input import regenerate_thinkers, SUBJECTS

    body = request.get_json() or {}
    subject_key = body.get("subject")  # optional — None means all subjects

    if subject_key and subject_key not in SUBJECTS:
        return jsonify({"error": f"Unknown subject: {subject_key}"}), 400

    # Run in background thread so the request returns immediately
    task_id = f"thinkers-{int(time.time())}"

    def _run():
        try:
            result = regenerate_thinkers(subject_key)
            # Store result so frontend can poll
            _thinker_tasks[task_id] = {"status": "done", **result}
        except Exception as e:
            _thinker_tasks[task_id] = {"status": "error", "error": str(e)}

    _thinker_tasks[task_id] = {"status": "running"}
    t = threading.Thread(target=_run, daemon=True)
    t.start()

    return jsonify({"task_id": task_id, "status": "running"})


# In-memory task tracker for thinker regeneration
_thinker_tasks = {}


@app.route("/api/reference-docs/update-thinkers/<task_id>")
def api_update_thinkers_status(task_id):
    """Poll the status of a thinker regeneration task."""
    task = _thinker_tasks.get(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404
    return jsonify(task)


# ========================================================================
# Scheme of Work Engine API (Prompt Sheet 13)
# ========================================================================

@app.route("/api/schemes")
def api_list_schemes():
    from sow_engine import list_schemes
    subject = request.args.get("subject")
    key_stage = request.args.get("key_stage")
    schemes = list_schemes()
    # Optional filtering
    if subject:
        schemes = [s for s in schemes if s["subject"].lower() == subject.lower()]
    if key_stage:
        schemes = [s for s in schemes if s["key_stage"].lower() == key_stage.lower()]
    return jsonify({"schemes": schemes})


@app.route("/api/schemes/<scheme_id>")
def api_get_scheme(scheme_id):
    from sow_engine import get_scheme
    scheme = get_scheme(scheme_id)
    if not scheme:
        return jsonify({"error": "Scheme not found"}), 404
    return jsonify(scheme)


@app.route("/api/schemes/<scheme_id>", methods=["PUT"])
def api_update_scheme(scheme_id):
    from sow_engine import get_scheme, save_scheme
    scheme = get_scheme(scheme_id)
    if not scheme:
        return jsonify({"error": "Scheme not found"}), 404
    body = request.get_json() or {}
    scheme.update(body)
    scheme["id"] = scheme_id
    saved = save_scheme(scheme)
    return jsonify({"saved": True, "scheme": saved})


@app.route("/api/schemes/<scheme_id>", methods=["DELETE"])
def api_delete_scheme(scheme_id):
    from sow_engine import delete_scheme
    delete_scheme(scheme_id)
    return jsonify({"deleted": True})


@app.route("/api/schemes/import", methods=["POST"])
def api_import_scheme():
    from access_tiers import get_tier, TIER_HIERARCHY
    if TIER_HIERARCHY.get(get_tier(), 0) < TIER_HIERARCHY["all_custom"]:
        return jsonify({"error": "upgrade_required", "message": "Scheme import requires All Custom access."}), 403
    from sow_engine import import_from_course
    body = request.get_json() or {}
    course_id = body.get("course_id", _active_course_id)
    config = courses.get_course(course_id)
    if not config:
        return jsonify({"error": "No course loaded. Please select a course first."}), 404
    try:
        data = get_data(course_id)
        if not data or not data.get("all_lessons"):
            return jsonify({"error": "No lesson data found for this course. Check the spreadsheet has been parsed correctly."}), 400
        scheme = import_from_course(config, data)
        scheme_path = str(Path(__file__).parent / "data" / "schemes" / f"{scheme['id']}.json")
        return jsonify({"saved": True, "scheme": scheme, "file_path": scheme_path})
    except Exception as e:
        logger.error(f"Scheme import failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/schemes/upload", methods=["POST"])
def api_upload_scheme():
    """Upload a scheme file (xlsx, csv) and parse it into a scheme."""
    from sow_engine import import_from_file
    import tempfile

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400

    ext = Path(f.filename).suffix.lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        return jsonify({"error": f"Unsupported file type: {ext}. Upload .xlsx, .xls, or .csv"}), 400

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    f.save(tmp.name)
    tmp.close()

    try:
        scheme = import_from_file(
            file_path=tmp.name,
            subject=request.form.get("subject", ""),
            key_stage=request.form.get("key_stage", ""),
            year_group=request.form.get("year_group", ""),
            exam_board=request.form.get("exam_board", ""),
        )
        scheme_path = str(Path(__file__).parent / "data" / "schemes" / f"{scheme['id']}.json")
        return jsonify({"saved": True, "scheme": scheme, "file_path": scheme_path})
    except Exception as e:
        logger.error(f"Scheme file upload failed: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp.name)


@app.route("/api/schemes/<scheme_id>/save-to-drive", methods=["POST"])
def api_scheme_save_to_drive(scheme_id):
    """Export a scheme as Excel and upload to Google Drive."""
    from sow_engine import get_scheme, export_to_excel
    import tempfile

    scheme = get_scheme(scheme_id)
    if not scheme:
        return jsonify({"error": "Scheme not found"}), 404

    try:
        # Generate Excel file
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.close()
        export_to_excel(scheme_id, output_path=tmp.name)

        # Upload to Google Drive as Google Sheets
        from gdrive import upload_as_google_native
        safe_title = re.sub(r"[^\w\s-]", "", scheme.get("title", "Scheme")).strip()
        result = upload_as_google_native(
            local_path=tmp.name,
            name=f"{safe_title} - Scheme of Work",
            target_mime="application/vnd.google-apps.spreadsheet",
        )
        os.unlink(tmp.name)
        return jsonify({"saved": True, "file_id": result["file_id"], "web_link": result["web_link"]})
    except Exception as e:
        logger.error(f"Scheme save to Drive failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/schemes/generate", methods=["POST"])
def api_generate_scheme():
    from access_tiers import get_tier, TIER_HIERARCHY
    if TIER_HIERARCHY.get(get_tier(), 0) < TIER_HIERARCHY["all_custom"]:
        return jsonify({"error": "upgrade_required", "message": "Scheme generation requires All Custom access."}), 403
    from sow_engine import generate_scheme
    import reference_library as reflib

    body = request.get_json() or {}
    subject = body.get("subject")
    key_stage = body.get("key_stage")
    scope = body.get("scope", "single_year")
    year_group = body.get("year_group")
    if not subject or not key_stage:
        return jsonify({"error": "subject and key_stage are required"}), 400

    # Determine which year groups to generate for
    KS_YEAR_GROUPS = {
        "KS1": [1, 2],
        "KS2": [3, 4, 5, 6],
        "KS3": [7, 8, 9],
        "KS4": [10, 11],
        "KS5": [12, 13],
    }
    if scope == "whole_ks":
        year_groups = KS_YEAR_GROUPS.get(key_stage, [])
        if not year_groups:
            return jsonify({"error": f"Unknown key stage: {key_stage}"}), 400
    else:
        if not year_group:
            return jsonify({"error": "year_group is required for single year generation"}), 400
        year_groups = [int(year_group)]

    ref_context = reflib.get_reference_context(
        subject=subject, key_stage=key_stage,
        exam_board=body.get("exam_board"),
    )

    try:
        schemes = []
        for yg in year_groups:
            scheme = generate_scheme(
                subject=subject,
                key_stage=key_stage,
                year_group=yg,
                lessons_per_week=int(body.get("lessons_per_week", 3)),
                weeks_per_term=int(body.get("weeks_per_term", 6)),
                exam_board=body.get("exam_board"),
                priorities=body.get("priorities"),
                exclusions=body.get("exclusions"),
                reference_context=ref_context,
                model=body.get("model", "claude-sonnet-4-5-20250929"),
            )
            schemes.append(scheme)
        if len(schemes) == 1:
            return jsonify({"saved": True, "scheme": schemes[0]})
        else:
            return jsonify({
                "saved": True,
                "scheme": schemes[0],
                "all_schemes": schemes,
                "message": f"Generated {len(schemes)} schemes for {key_stage} (Years {', '.join(str(y) for y in year_groups)})"
            })
    except Exception as e:
        logger.error(f"Scheme generation failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/schemes/<scheme_id>/review", methods=["POST"])
def api_review_scheme(scheme_id):
    from sow_engine import review_scheme, get_scheme
    import reference_library as reflib

    scheme = get_scheme(scheme_id)
    if not scheme:
        return jsonify({"error": "Scheme not found"}), 404

    body = request.get_json() or {}
    ref_context = reflib.get_reference_context(
        subject=scheme.get("subject"),
        key_stage=scheme.get("keyStage"),
        exam_board=scheme.get("examBoard"),
    )

    try:
        review = review_scheme(
            scheme_id,
            reference_context=ref_context,
            model=body.get("model", "claude-sonnet-4-5-20250929"),
        )
        return jsonify({"review": review})
    except Exception as e:
        logger.error(f"Scheme review failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/schemes/<scheme_id>/review", methods=["GET"])
def api_get_review(scheme_id):
    from sow_engine import get_review
    review = get_review(scheme_id)
    if not review:
        return jsonify({"error": "No review found"}), 404
    return jsonify({"review": review})


@app.route("/api/schemes/<scheme_id>/apply-suggestion", methods=["POST"])
def api_apply_suggestion(scheme_id):
    from sow_engine import apply_suggestion
    import reference_library as reflib
    from sow_engine import get_scheme

    scheme = get_scheme(scheme_id)
    if not scheme:
        return jsonify({"error": "Scheme not found"}), 404

    body = request.get_json() or {}
    suggestion = body.get("suggestion")
    if not suggestion:
        return jsonify({"error": "suggestion is required"}), 400

    ref_context = reflib.get_reference_context(
        subject=scheme.get("subject"),
        key_stage=scheme.get("keyStage"),
    )

    try:
        updated = apply_suggestion(scheme_id, suggestion, reference_context=ref_context)
        return jsonify({"saved": True, "scheme": updated})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/schemes/<scheme_id>/export-excel", methods=["GET", "POST"])
def api_export_scheme_excel(scheme_id):
    from sow_engine import export_to_excel
    try:
        path = export_to_excel(scheme_id)
        return send_file(path, as_attachment=True)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


# ========================================================================
# Progression Map API (Prompt Sheet 14)
# ========================================================================

@app.route("/api/progression-maps/generate", methods=["POST"])
def api_generate_progression_map():
    from progression_map import generate_progression_map
    from sow_engine import get_scheme
    import reference_library as reflib

    body = request.get_json() or {}
    scheme_id = body.get("scheme_id")
    if not scheme_id:
        return jsonify({"error": "scheme_id is required"}), 400

    scheme = get_scheme(scheme_id)
    if not scheme:
        return jsonify({"error": "Scheme not found"}), 404

    ref_context = reflib.get_reference_context(
        subject=scheme.get("subject"),
        key_stage=scheme.get("keyStage"),
    )

    try:
        map_data = generate_progression_map(
            scheme, reference_context=ref_context,
            model=body.get("model", "claude-sonnet-4-5-20250929"),
        )
        return jsonify({"saved": True, "map": map_data})
    except Exception as e:
        logger.error(f"Progression map generation failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/progression-maps")
def api_list_progression_maps():
    from progression_map import list_maps
    subject = request.args.get("subject")
    return jsonify({"maps": list_maps(subject=subject)})


@app.route("/api/progression-maps/<map_id>")
def api_get_progression_map(map_id):
    from progression_map import get_map
    m = get_map(map_id)
    if not m:
        return jsonify({"error": "Map not found"}), 404
    return jsonify(m)


@app.route("/api/progression-maps/<map_id>", methods=["PUT"])
def api_update_progression_map(map_id):
    from progression_map import update_map
    body = request.get_json() or {}
    try:
        m = update_map(map_id, body)
        return jsonify({"updated": True, "map": m})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/progression-maps/<map_id>", methods=["DELETE"])
def api_delete_progression_map(map_id):
    from progression_map import delete_map
    delete_map(map_id)
    return jsonify({"deleted": True})


@app.route("/api/progression-maps/<map_id>/svg")
def api_progression_map_svg(map_id):
    from progression_map import get_map, render_svg
    m = get_map(map_id)
    if not m:
        return jsonify({"error": "Map not found"}), 404
    svg = render_svg(m)
    return Response(svg, mimetype="image/svg+xml")


@app.route("/api/progression-maps/<map_id>/download/<fmt>")
def api_progression_map_download(map_id, fmt):
    """Download a progression map as SVG or PDF."""
    from progression_map import get_map
    m = get_map(map_id)
    if not m:
        return jsonify({"error": "Map not found"}), 404

    exported = m.get("exported_files", {})
    if fmt == "svg":
        path = exported.get("svg")
        if not path or not Path(path).exists():
            return jsonify({"error": "SVG file not found. Try regenerating the map."}), 404
        title = m.get("title", "Progression Map")
        return send_file(path, as_attachment=True, download_name=f"{title}.svg")
    elif fmt == "pdf":
        path = exported.get("pdf")
        if not path or not Path(path).exists():
            return jsonify({"error": "PDF not available. LibreOffice may not be installed."}), 404
        title = m.get("title", "Progression Map")
        return send_file(path, as_attachment=True, download_name=f"{title}.pdf")
    else:
        return jsonify({"error": f"Unknown format: {fmt}"}), 400


@app.route("/api/progression-maps/<map_id>/save-to-drive", methods=["POST"])
def api_progression_map_save_to_drive(map_id):
    """Upload a progression map SVG (or PDF) to Google Drive."""
    from progression_map import get_map
    from gdrive import upload_as_google_native

    m = get_map(map_id)
    if not m:
        return jsonify({"error": "Map not found"}), 404

    exported = m.get("exported_files", {})
    title = m.get("title", "Progression Map")

    config = get_active_course_config()
    folder_id = (config or {}).get("gdrive_folder_id") or None

    # Prefer PDF for Drive upload, fall back to SVG
    file_path = exported.get("pdf") or exported.get("svg")
    if not file_path or not Path(file_path).exists():
        return jsonify({"error": "No exported file found. Try regenerating the map."}), 404

    ext = Path(file_path).suffix.lower()
    try:
        if ext == ".pdf":
            # Upload PDF directly (no conversion needed)
            from googleapiclient.http import MediaFileUpload
            from gdrive import _get_service, MIME_TYPES
            service = _get_service()
            metadata = {"name": f"{title}.pdf"}
            if folder_id:
                metadata["parents"] = [folder_id]
            media = MediaFileUpload(file_path, mimetype="application/pdf")
            result = service.files().create(
                body=metadata, media_body=media,
                fields="id, webViewLink",
            ).execute()
            return jsonify({
                "file_id": result["id"],
                "web_link": result.get("webViewLink", ""),
            })
        else:
            # SVG — upload as raw file
            from googleapiclient.http import MediaFileUpload
            from gdrive import _get_service
            service = _get_service()
            metadata = {"name": f"{title}.svg"}
            if folder_id:
                metadata["parents"] = [folder_id]
            media = MediaFileUpload(file_path, mimetype="image/svg+xml")
            result = service.files().create(
                body=metadata, media_body=media,
                fields="id, webViewLink",
            ).execute()
            return jsonify({
                "file_id": result["id"],
                "web_link": result.get("webViewLink", ""),
            })
    except Exception as e:
        logger.error(f"Drive upload failed for progression map {map_id}: {e}")
        return jsonify({"error": str(e)}), 500


# ========================================================================
# Booklet Types API (Prompt Sheet 15)
# ========================================================================

@app.route("/api/booklet-types/generate", methods=["POST"])
def api_generate_typed_booklet():
    from booklet_types import generate_typed_booklet
    from generator import check_existing_booklet
    import reference_library as reflib

    body = request.get_json() or {}
    year = body.get("year")
    lesson_num = body.get("lesson_num")
    booklet_type = body.get("type")

    if not all([year, lesson_num, booklet_type]):
        return jsonify({"error": "year, lesson_num, and type are required"}), 400

    data = get_data()
    lesson = _find_lesson(data, int(year), int(lesson_num))
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    config = get_active_course_config()

    # Try to load existing booklet content
    existing = check_existing_booklet(lesson)
    existing_content = None
    if existing["exists"]:
        md_path = existing["docx_path"].replace(".docx", ".md")
        md_file = Path(md_path)
        if md_file.exists():
            existing_content = md_file.read_text(encoding="utf-8")

    # Get reference context
    ref_context = reflib.get_reference_context(
        subject=lesson.get("subject"),
        key_stage=f"KS{config.get('key_stage', '')}" if config.get("key_stage") else None,
        exam_board=config.get("exam_board"),
    )

    # For revision, get all lessons in the same topic/unit
    unit_lessons = None
    if booklet_type == "revision":
        topic = lesson.get("topic")
        if topic:
            unit_lessons = [
                l for l in data["booklet_lessons"]
                if l.get("topic") == topic and l.get("year") == lesson.get("year")
            ]

    try:
        result = generate_typed_booklet(
            lesson=lesson,
            booklet_type=booklet_type,
            course_config=config,
            existing_booklet_content=existing_content,
            reference_context=ref_context,
            unit_lessons=unit_lessons,
            model=body.get("model", "claude-sonnet-4-5-20250929"),
        )
        return jsonify({"success": True, **result})
    except Exception as e:
        logger.error(f"Typed booklet generation failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/booklet-types/list/<int:year>/<int:lesson_num>")
def api_list_typed_booklets(year, lesson_num):
    from booklet_types import list_typed_booklets
    data = get_data()
    lesson = _find_lesson(data, year, lesson_num)
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404
    booklets = list_typed_booklets(lesson)
    return jsonify({"booklets": booklets})


# ========================================================================
# Presentation Engine API (Prompt Sheet 16)
# ========================================================================

@app.route("/api/presentations/generate", methods=["POST"])
def api_generate_presentation():
    from presentation_engine import generate_slide_content, render_pptx, save_presentation_data
    from generator import check_existing_booklet
    import reference_library as reflib

    body = request.get_json() or {}
    year = body.get("year")
    lesson_num = body.get("lesson_num")

    if not all([year, lesson_num]):
        return jsonify({"error": "year and lesson_num are required"}), 400

    data = get_data()
    lesson = _find_lesson(data, int(year), int(lesson_num))
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    config = get_active_course_config()

    # Load existing booklet if available
    existing = check_existing_booklet(lesson)
    existing_content = None
    if existing["exists"]:
        md_path = existing["docx_path"].replace(".docx", ".md")
        md_file = Path(md_path)
        if md_file.exists():
            existing_content = md_file.read_text(encoding="utf-8")

    ref_context = reflib.get_reference_context(
        subject=lesson.get("subject"),
        key_stage=f"KS{config.get('key_stage', '')}" if config.get("key_stage") else None,
    )

    try:
        slide_result = generate_slide_content(
            lesson=lesson,
            reference_context=ref_context,
            existing_booklet_content=existing_content,
            include_speaker_notes=body.get("include_speaker_notes", True),
            include_differentiation=body.get("include_differentiation", True),
            model=body.get("model", "claude-sonnet-4-5-20250929"),
        )

        pptx_path = render_pptx(
            slides=slide_result["slides"],
            lesson=lesson,
            style=body.get("style", "clean"),
        )

        pres_data = save_presentation_data(
            lesson=lesson,
            slides=slide_result["slides"],
            pptx_path=pptx_path,
            usage=slide_result["usage"],
            model=slide_result["model"],
            duration_s=slide_result["duration_s"],
        )

        return jsonify({"success": True, "presentation": pres_data})
    except Exception as e:
        logger.error(f"Presentation generation failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/presentations")
def api_list_presentations():
    from presentation_engine import list_presentations
    return jsonify({"presentations": list_presentations()})


@app.route("/api/presentations/<pres_id>")
def api_get_presentation(pres_id):
    from presentation_engine import get_presentation
    p = get_presentation(pres_id)
    if not p:
        return jsonify({"error": "Presentation not found"}), 404
    return jsonify(p)


@app.route("/api/presentations/<pres_id>/download")
def api_download_presentation(pres_id):
    from presentation_engine import get_presentation
    p = get_presentation(pres_id)
    if not p:
        return jsonify({"error": "Presentation not found"}), 404
    pptx_path = p.get("pptx_path")
    if not pptx_path or not Path(pptx_path).exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(pptx_path, as_attachment=True)


@app.route("/api/presentations/gamma", methods=["POST"])
def api_export_gamma():
    from presentation_engine import generate_slide_content, export_gamma_format
    import reference_library as reflib

    body = request.get_json() or {}
    year = body.get("year")
    lesson_num = body.get("lesson_num")
    if not all([year, lesson_num]):
        return jsonify({"error": "year and lesson_num required"}), 400

    data = get_data()
    lesson = _find_lesson(data, int(year), int(lesson_num))
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404

    ref_context = reflib.get_reference_context(subject=lesson.get("subject"))

    slide_result = generate_slide_content(lesson=lesson, reference_context=ref_context)
    gamma_text = export_gamma_format(slide_result["slides"], lesson)
    return jsonify({"gamma_text": gamma_text, "slide_count": len(slide_result["slides"])})


# ========================================================================
# Assessment Builder API (Prompt Sheet 17)
# ========================================================================

@app.route("/api/assessments/generate", methods=["POST"])
def api_generate_assessment():
    from assessment_engine import generate_assessment
    import reference_library as reflib

    body = request.get_json() or {}
    lesson_refs = body.get("lessons", [])  # [{year, lesson_num}, ...]

    if not lesson_refs:
        return jsonify({"error": "lessons list is required"}), 400

    data = get_data()
    config = get_active_course_config()

    # Resolve lesson refs to lesson objects
    lessons = []
    for ref in lesson_refs:
        l = _find_lesson(data, int(ref.get("year", 0)), int(ref.get("lesson_num", 0)))
        if l:
            lessons.append(l)

    if not lessons:
        return jsonify({"error": "No matching lessons found"}), 404

    subject = lessons[0].get("subject", "Unknown")
    year_group = lessons[0].get("year", 0)

    ref_context = reflib.get_reference_context(
        subject=subject,
        key_stage=f"KS{config.get('key_stage', '')}" if config.get("key_stage") else None,
        exam_board=config.get("exam_board"),
    )

    try:
        assessment = generate_assessment(
            lessons=lessons,
            subject=subject,
            year_group=year_group,
            config=body.get("config", {}),
            num_mc=int(body.get("num_mc", 10)),
            num_short=int(body.get("num_short", 8)),
            num_long=int(body.get("num_long", 3)),
            difficulty=body.get("difficulty", "medium"),
            assessment_type=body.get("assessment_type", "custom"),
            reference_context=ref_context,
            model=body.get("model", "claude-sonnet-4-5-20250929"),
        )
        return jsonify({"success": True, "assessment": assessment})
    except Exception as e:
        logger.error(f"Assessment generation failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/assessments")
def api_list_assessments():
    from assessment_engine import list_assessments
    subject = request.args.get("subject")
    return jsonify({"assessments": list_assessments(subject=subject)})


@app.route("/api/assessments/<assessment_id>")
def api_get_assessment(assessment_id):
    from assessment_engine import get_assessment
    a = get_assessment(assessment_id)
    if not a:
        return jsonify({"error": "Assessment not found"}), 404
    return jsonify(a)


@app.route("/api/assessments/<assessment_id>", methods=["PUT"])
def api_update_assessment(assessment_id):
    from assessment_engine import update_assessment
    body = request.get_json() or {}
    try:
        a = update_assessment(assessment_id, body)
        return jsonify({"updated": True, "assessment": a})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/assessments/<assessment_id>", methods=["DELETE"])
def api_delete_assessment(assessment_id):
    from assessment_engine import delete_assessment
    delete_assessment(assessment_id)
    return jsonify({"deleted": True})


@app.route("/api/assessments/<assessment_id>/export/student-paper")
def api_export_student_paper(assessment_id):
    from assessment_engine import export_student_paper
    try:
        path = export_student_paper(assessment_id)
        return send_file(path, as_attachment=True)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


@app.route("/api/assessments/<assessment_id>/export/mark-scheme")
def api_export_mark_scheme(assessment_id):
    from assessment_engine import export_mark_scheme
    try:
        path = export_mark_scheme(assessment_id)
        return send_file(path, as_attachment=True)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


# Question bank
@app.route("/api/question-bank")
def api_question_bank():
    from assessment_engine import list_question_bank
    return jsonify({"questions": list_question_bank(
        subject=request.args.get("subject"),
        topic=request.args.get("topic"),
        question_type=request.args.get("type"),
        difficulty=request.args.get("difficulty"),
        starred_only=request.args.get("starred") == "true",
    )})


@app.route("/api/question-bank/<question_id>/star", methods=["PUT"])
def api_star_question(question_id):
    from assessment_engine import star_question
    body = request.get_json() or {}
    starred = body.get("starred", True)
    try:
        q = star_question(question_id, starred)
        return jsonify({"starred": q["is_starred"]})
    except ValueError as e:
        return jsonify({"error": str(e)}), 404


# ========================================================================
# Helpers
# ========================================================================

def _find_lesson(data, year, lesson_num):
    """Find a booklet lesson by year and number."""
    for l in data["booklet_lessons"]:
        if l["year"] == year and l["lesson_number"] == lesson_num:
            return l
    return None


# ─── File preview & open folder ─────────────────────────────────────────
import subprocess as _subprocess

@app.route("/api/preview-file")
def api_preview_file():
    """Serve a generated file for preview in the browser."""
    file_path = request.args.get("path", "")
    if not file_path:
        return "No path", 400
    p = Path(file_path)
    if not p.exists():
        return f"File not found: {file_path}", 404
    # Only serve files under output/ or data/ for safety
    try:
        p.resolve().relative_to(Path(__file__).parent.resolve())
    except ValueError:
        return "Access denied", 403
    return send_file(str(p.resolve()), as_attachment=False)


@app.route("/api/open-folder", methods=["POST"])
def api_open_folder():
    """Open the containing folder of a file in Finder/Explorer."""
    body = request.get_json() or {}
    file_path = body.get("path", "")
    if not file_path:
        return jsonify({"error": "No path"}), 400
    p = Path(file_path)
    folder = p.parent if p.is_file() else p
    if not folder.exists():
        return jsonify({"error": "Folder not found"}), 404
    try:
        import platform
        if platform.system() == "Darwin":
            _subprocess.Popen(["open", str(folder)])
        elif platform.system() == "Windows":
            _subprocess.Popen(["explorer", str(folder)])
        else:
            _subprocess.Popen(["xdg-open", str(folder)])
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5050)
